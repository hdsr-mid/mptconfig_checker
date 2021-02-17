from mptconfig import constants
from mptconfig.fews_utilities import FewsConfig
from mptconfig.fews_utilities import xml_to_dict
from mptconfig.utils import get_validation_attribs
from mptconfig.utils import idmap2tags
from mptconfig.utils import sort_validation_attribs
from mptconfig.utils import update_date
from mptconfig.utils import update_histtag
from mptconfig.utils import update_hlocs
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from pathlib import Path
from shapely.geometry import Point  # noqa shapely comes with geopandas
from typing import Dict
from typing import List
from typing import Optional
from typing import Tuple

import logging
import numpy as np  # noqa numpy comes with geopandas
import pandas as pd  # noqa pandas comes with geopandas
import re
import shutil


logger = logging.getLogger(__name__)

pd.options.mode.chained_assignment = None

constants.check_constants()


class MptConfigResults(dict):
    def __init__(self):
        super().__init__()
        """Read input-excel (consistency_input_xlsx) into a dictionary with dataframes, each
        dataframe representing an excel sheet."""
        _dict = pd.read_excel(
            io=constants.PathConstants.consistency_input_xlsx.path, sheet_name=None, engine="openpyxl"
        )
        for sheet_name, sheet_df in _dict.items():
            if sheet_name not in constants.FIXED_SHEETS:
                continue
            super().__setitem__(sheet_name, sheet_df)

    def __getitem__(self, sheet_name: str) -> pd.DataFrame:
        if sheet_name not in self:
            raise AttributeError(f"No such sheet_name: {sheet_name}")
        return super().__getitem__(sheet_name)

    def __setitem__(self, sheet_name: str, sheet_df: pd.DataFrame) -> None:
        if sheet_name in self:
            raise AssertionError(f"sheet_name {sheet_name} already exists")
        if not isinstance(sheet_name, str):
            raise AssertionError(f"sheet_name {sheet_name} must be a string and not a {type(sheet_name)}")
        if not isinstance(sheet_df, pd.DataFrame):
            raise AssertionError(f"sheet_df must be a panda Dataframe and not a {type(sheet_df)}")
        super().__setitem__(sheet_name, sheet_df)

    def __delitem__(self, sheet_name: str) -> None:
        if sheet_name not in self:
            raise AttributeError(f"No such sheet_name: {sheet_name}")
        super().__delitem__(sheet_name)


class MptConfigChecker:
    """Class to read, check and write a HDSR meetpuntconfiguratie.
    The main property of the class is 'self.consistency' which is:
        - a dictionary of panda dataframes: each dataframe is one excel sheet.
        - updated throughout the whole class with check results
    """

    def __init__(self):
        self._location_sets = None
        self._histtags = None
        self._hoofdloc = None
        self._hoofdloc_new = None
        self._subloc = None
        self._waterstandloc = None
        self._mswloc = None
        self._mpthisttags = None
        self._fews_config = None
        self._ignored_exloc = None
        self._ignored_histtag = None
        self._ignored_ts800 = None
        self._ignored_xy = None
        self.consistency = MptConfigResults()

    @property
    def fews_config(self):
        if self._fews_config is not None:
            return self._fews_config
        self._fews_config = FewsConfig(path=constants.PathConstants.fews_config.path)
        return self._fews_config

    @property
    def location_sets(self) -> Dict:
        if self._location_sets is not None:
            return self._location_sets
        self._location_sets = {}
        for key, value in constants.LOCATIONS_SETS.items():
            if value not in self.fews_config.location_sets.keys():
                raise AssertionError(f"locationSet {key} specified in constants.LOCATIONS_SETS not in fews-config")
            elif "csvFile" in self.fews_config.location_sets[value].keys():
                self._location_sets[key] = {
                    "id": value,
                    "gdf": self.fews_config.get_locations(location_set_key=value),
                }
            else:
                raise AssertionError(f"{key} not a csvFile location-set")
        return self._location_sets

    @property
    def histtags(self) -> pd.DataFrame:
        if self._histtags is not None:
            return self._histtags
        logger.info(f"reading histags: {constants.PathConstants.histtags_csv.path}")
        dtype_columns = ["total_min_start_dt", "total_max_end_dt"]
        self._histtags = pd.read_csv(
            filepath_or_buffer=constants.PathConstants.histtags_csv.path,
            parse_dates=dtype_columns,
            sep=None,
            engine="python",
        )

        for dtype_column in dtype_columns:
            if not pd.api.types.is_datetime64_dtype(self.histtags[dtype_column]):
                raise AssertionError(
                    f"dtype_column {dtype_column} in {constants.PathConstants.histtags_csv.path} can not be converted "
                    f"to np.datetime64. Check if values are dates."
                )
        return self._histtags

    @property
    def mswloc(self) -> pd.DataFrame:
        if self._mswloc is not None:
            return self._mswloc
        self._mswloc = self.fews_config.get_locations(location_set_key="MSW_STATIONS")
        return self._mswloc

    @property
    def waterstandloc(self) -> pd.DataFrame:
        if self._waterstandloc is not None:
            return self._waterstandloc
        self._waterstandloc = self.fews_config.get_locations(location_set_key="OPVLWATER_WATERSTANDEN_AUTO")
        return self._waterstandloc

    @property
    def subloc(self) -> pd.DataFrame:
        if self._subloc is not None:
            return self._subloc
        self._subloc = self.fews_config.get_locations(location_set_key="OPVLWATER_SUBLOC")
        return self._subloc

    @property
    def hoofdloc(self) -> pd.DataFrame:
        """Get the latest version of the hoofdlocation: hoofdloc_new or else hoofdloc"""
        if self.hoofdloc_new is not None:
            return self.hoofdloc_new
        if self._hoofdloc is not None:
            return self._hoofdloc
        self._hoofdloc = self.fews_config.get_locations(location_set_key="OPVLWATER_HOOFDLOC")
        return self._hoofdloc

    @property
    def hoofdloc_new(self) -> Optional[pd.DataFrame]:
        return self._hoofdloc_new

    @property
    def mpt_histtags(self) -> pd.DataFrame:
        if self._mpthisttags is not None:
            return self._mpthisttags
        idmaps = self._get_idmaps()
        histtags_df = self.histtags.copy()
        histtags_df["fews_locid"] = histtags_df.apply(func=idmap2tags, args=[idmaps], axis=1)
        histtags_df = histtags_df[histtags_df["fews_locid"].notna()]
        self._mpthisttags = histtags_df.explode("fews_locid").reset_index(drop=True)
        return self._mpthisttags

    @property
    def ignored_exloc(self) -> pd.DataFrame:
        if self._ignored_exloc is not None:
            return self._ignored_exloc
        logger.info(f"reading {constants.PathConstants.ignored_exloc.path}")
        self._ignored_exloc = pd.read_csv(
            filepath_or_buffer=constants.PathConstants.ignored_exloc.path, sep=",", engine="python",
        )
        return self._ignored_exloc

    @property
    def ignored_histtag(self) -> pd.DataFrame:
        if self._ignored_histtag is not None:
            return self._ignored_histtag
        logger.info(f"reading {constants.PathConstants.ignored_histtag.path}")
        self._ignored_histtag = pd.read_csv(
            filepath_or_buffer=constants.PathConstants.ignored_histtag.path, sep=",", engine="python",
        )
        self._ignored_histtag["UNKNOWN_SERIE"] = self._ignored_histtag["UNKNOWN_SERIE"].str.replace("#", "")
        return self._ignored_histtag

    @property
    def ignored_ts800(self) -> pd.DataFrame:
        if self._ignored_ts800 is not None:
            return self._ignored_ts800
        logger.info(f"reading {constants.PathConstants.ignored_ts800.path}")
        self._ignored_ts800 = pd.read_csv(
            filepath_or_buffer=constants.PathConstants.ignored_ts800.path, sep=",", engine="python",
        )
        return self._ignored_ts800

    @property
    def ignored_xy(self) -> pd.DataFrame:
        if self._ignored_xy is not None:
            return self._ignored_xy
        logger.info(f"reading {constants.PathConstants.ignored_xy.path}")
        self._ignored_xy = pd.read_csv(
            filepath_or_buffer=constants.PathConstants.ignored_xy.path, sep=",", engine="python",
        )
        return self._ignored_xy

    def _get_idmaps(self, idmap_files: List[str] = None) -> List[Dict]:
        """Get id mapping from 1 or more sources (xml files) and return them in a flatted list.
        Example:
            idmap_files = ['IdOPVLWATER', 'IdOPVLWATER_HYMOS', 'IdHDSR_NSC', 'IdOPVLWATER_WQ', 'IdGrondwaterCAW']
        returns:
            [
                ...,
                {'externalLocation': '7612', 'externalParameter': 'HB1', 'internalLocation': 'OW761202', 'internalParameter': 'H.G.0'},  # noqa
                {'externalLocation': '7612', 'externalParameter': 'HO1', 'internalLocation': 'OW761201', 'internalParameter': 'H.G.0'},  # noqa
                ...,
            ]
        """
        if not idmap_files:
            idmap_files = constants.IDMAP_FILES
        idmaps = [xml_to_dict(xml_filepath=self.fews_config.IdMapFiles[idmap])["idMap"]["map"] for idmap in idmap_files]
        return [item for sublist in idmaps for item in sublist]

    def _create_hoofdloc_new(self, par_dict: Dict) -> None:
        """Create a new hoofdloc from sublocs in case no errors found during
        check_hloc_consistency in case all sublocs of same hloc have consistent parameters."""
        assert isinstance(par_dict, dict), f"par_dict should be a dictionary, not a {type(par_dict)}"
        par_gdf = pd.DataFrame(par_dict)
        columns = list(self.hoofdloc.columns)
        drop_cols = [col for col in self.hoofdloc.columns if col in par_gdf.columns and col != "LOC_ID"]
        drop_cols = drop_cols + ["geometry"]
        self._hoofdloc_new = self.hoofdloc.drop(drop_cols, axis=1, inplace=False)
        self._hoofdloc_new = par_gdf.merge(self._hoofdloc_new, on="LOC_ID")
        self._hoofdloc_new["geometry"] = self._hoofdloc_new.apply(
            func=(lambda x: Point(float(x["X"]), float(x["Y"]))), axis=1
        )
        self._hoofdloc_new = self._hoofdloc_new[columns]

    def _update_staff_gauge(self, row: pd.Series) -> Tuple[str, str]:
        """Assign upstream and downstream staff gauges to subloc."""
        result = {"HBOV": "", "HBEN": ""}
        for key in result.keys():
            df = self.waterstandloc.loc[self.waterstandloc["LOC_ID"] == row[key]]
            if not df.empty:
                result[key] = df["PEILSCHAAL"].values[0]
        return result["HBOV"], result["HBEN"]

    def histtags_to_mpt(self, sheet_name: str = "mpt") -> Tuple[str, pd.DataFrame]:
        """Convert histTag-ids to mpt-ids."""
        mpt_df = pd.concat(
            [
                self.mpt_histtags.groupby(["fews_locid"], sort=False)["total_min_start_dt"].min(),
                self.mpt_histtags.groupby(["fews_locid"], sort=False)["total_max_end_dt"].max(),
            ],
            axis=1,
        )

        mpt_df = mpt_df.sort_index(axis=0)
        mpt_df.columns = ["STARTDATE", "ENDDATE"]
        mpt_df.index.name = "LOC_ID"
        kw_locs = list(mpt_df[mpt_df.index.str.contains("KW", regex=False)].index)
        h_locs = np.unique([f"{loc[0:-1]}0" for loc in kw_locs])
        h_locs_missing = [loc for loc in h_locs if loc not in list(mpt_df.index)]
        h_locs_df = pd.DataFrame(
            data={
                "LOC_ID": h_locs_missing,
                "STARTDATE": [pd.NaT] * len(h_locs_missing),
                "ENDDATE": [pd.NaT] * len(h_locs_missing),
            }
        )

        h_locs_df = h_locs_df.set_index("LOC_ID")
        mpt_df = pd.concat([mpt_df, h_locs_df], axis=0)
        mpt_df[["STARTDATE", "ENDDATE"]] = mpt_df.apply(
            func=update_hlocs, args=[h_locs, mpt_df], axis=1, result_type="expand"
        )
        mpt_df = mpt_df.sort_index()
        return sheet_name, mpt_df

    def check_idmap_sections(self, sheet_name: str = "idmap section error") -> Tuple[str, pd.DataFrame]:
        """Check if all KW/OW locations are in the correct section."""
        logger.info(f"start {self.check_idmap_sections.__name__}")
        result_df = pd.DataFrame(
            columns=["bestand", "externalLocation", "externalParameter", "internalLocation", "internalParameter"]
        )
        for idmap, subsecs in constants.IDMAP_SECTIONS.items():
            for section_type, sections in subsecs.items():
                for section in sections:

                    xml_filepath = self.fews_config.IdMapFiles[idmap]
                    # not all section have a 'section_start' and 'section_end' so use kwarg 'section'
                    _dict = xml_to_dict(xml_filepath=xml_filepath, **section)
                    idmapping = _dict["idMap"]["map"]

                    prefix = constants.SECTION_TYPE_PREFIX_MAPPER[section_type]
                    pattern = fr"{prefix}\d{{6}}$"
                    idmap_wrong_section = [
                        idmap
                        for idmap in idmapping
                        if not bool(re.match(pattern=pattern, string=idmap["internalLocation"]))
                    ]
                    if not idmap_wrong_section:
                        continue
                    section_start = section.get("section_start", "")
                    section_end = section.get("section_end", "")
                    logger.warning(
                        (
                            f"{len(idmap_wrong_section)} "
                            f"internalLocations not {prefix}XXXXXX "
                            f"between {section_start} and {section_end} "
                            f"in {idmap}."
                        )
                    )
                    df = pd.DataFrame(idmap_wrong_section)
                    df["sectie"] = section_start
                    df["bestand"] = idmap
                    result_df = pd.concat(objs=[result_df, df], axis=0, join="outer")
        return sheet_name, result_df

    def check_missing_histtags(self, sheet_name: str = "histTags noMatch") -> Tuple[str, pd.DataFrame]:
        """Check if hisTags are missing in config."""
        logger.info(f"start {self.check_missing_histtags.__name__}")

        histtags_df = self.histtags.copy()

        idmaps = self._get_idmaps()

        histtags_df["fews_locid"] = self.histtags.apply(func=idmap2tags, args=[idmaps], axis=1)

        histtags_no_match_df = histtags_df[histtags_df["fews_locid"].isna()]
        histtags_no_match_df = histtags_no_match_df[
            ~histtags_no_match_df["serie"].isin(values=self.ignored_histtag["UNKNOWN_SERIE"])
        ]

        histtags_no_match_df = histtags_no_match_df.drop("fews_locid", axis=1)
        histtags_no_match_df.columns = ["UNKNOWN_SERIE", "STARTDATE", "ENDDATE"]
        histtags_no_match_df = histtags_no_match_df.set_index("UNKNOWN_SERIE")
        if not histtags_no_match_df.empty:
            logger.warning(f"{len(histtags_no_match_df)} histTags not in idMaps")
        else:
            logger.info("all histTags in idMaps")
        return sheet_name, histtags_no_match_df

    def check_ignored_histtags(
        self, sheet_name: str = "ignored histTags match", idmap_files: List[str] = None,
    ) -> Tuple[str, pd.DataFrame]:
        """Check if ignored histTags do match with idmap."""
        logger.info(f"start {self.check_ignored_histtags.__name__}")
        assert isinstance(idmap_files, List) if idmap_files else True, "idmap_files must be a List"
        if not idmap_files:
            idmap_files = ["IdOPVLWATER"]
        histtags_opvlwater_df = self.histtags.copy()
        idmaps = self._get_idmaps(idmap_files=idmap_files)
        # TODO: @daniel kan fews_locid een lijst met meerdere loc_id's zijn?
        #  histtags_opvlwater_df.fews_locid.dtype is 'O' (is al string)
        histtags_opvlwater_df["fews_locid"] = self.histtags.apply(func=idmap2tags, args=[idmaps], axis=1)
        histtags_opvlwater_df = histtags_opvlwater_df[histtags_opvlwater_df["fews_locid"].notna()]
        histtag_ignore_match_df = self.ignored_histtag[
            self.ignored_histtag["UNKNOWN_SERIE"].isin(values=histtags_opvlwater_df["serie"])
        ]
        histtag_ignore_match_df = histtag_ignore_match_df.set_index("UNKNOWN_SERIE")
        if not histtag_ignore_match_df.empty:
            logger.warning(f"{len(histtag_ignore_match_df)} histTags should not be in histTags ignore")
        else:
            logger.info("hisTags ignore list consistent with idmaps")
        return sheet_name, histtag_ignore_match_df

    def check_double_idmaps(self, sheet_name: str = "idmaps double") -> Tuple[str, pd.DataFrame]:
        """Check if identical idmaps are doubled."""
        logger.info(f"start {self.check_double_idmaps.__name__}")
        result_df = pd.DataFrame(
            columns=["bestand", "externalLocation", "externalParameter", "internalLocation", "internalParameter"]
        )
        for idmap_file in constants.IDMAP_FILES:
            idmaps = self._get_idmaps(idmap_files=[idmap_file])
            idmap_doubles = [idmap for idmap in idmaps if idmaps.count(idmap) > 1]
            if not idmap_doubles:
                logger.info(f"No double idmaps in {idmap_file}")
                continue
            idmap_doubles = list({idmap["externalLocation"]: idmap for idmap in idmap_doubles}.values())
            logger.warning(f"{len(idmap_doubles)} double idmap(s) in {idmap_file}")
            df = pd.DataFrame(
                idmap_doubles,
                columns=["internalLocation", "externalLocation", "internalParameter", "externalParameter"],
            )
            df["bestand"] = idmap_file
            result_df = pd.concat(objs=[result_df, df], axis=0, join="outer")
        return sheet_name, result_df

    def check_missing_pars(self, sheet_name: str = "pars missing") -> Tuple[str, pd.DataFrame]:
        """Check if internal parameters in idmaps are missing in parameters.xml. """
        # TODO: @renier: put in docstring: alle id_mapping.xml inpars (bijv ‘H.R.0’) moeten voorkomen in RegionConfigFiles/parameters.xml
        logger.info(f"start {self.check_missing_pars.__name__}")
        config_parameters = list(self.fews_config.get_parameters(dict_keys="parameters").keys())

        idmaps = self._get_idmaps()
        id_map_parameters = [id_map["internalParameter"] for id_map in idmaps]
        params_missing = [parameter for parameter in id_map_parameters if parameter not in config_parameters]

        if len(params_missing) == 0:
            logger.info("all internal paramters are in config")
        else:
            logger.warning(f"{len(params_missing)} parameter(s) in idMaps are missing in config")
        result_df = pd.DataFrame({"parameters": params_missing})
        return sheet_name, result_df

    def check_hloc_consistency(self, sheet_name: str = "hloc error") -> Tuple[str, pd.DataFrame]:
        """Check if all sublocs of same hloc have consistent parameters."""
        logger.info(f"start {self.check_hloc_consistency.__name__}")
        xy_ignore_df = self.consistency.get("xy_ignore", pd.DataFrame({"internalLocation": [], "x": [], "y": []}))

        hloc_errors = {
            "LOC_ID": [],
            "SUB_LOCS": [],
            "LOC_NAME": [],
            "GEOMETRY": [],
            "SYSTEEM": [],
            "RAYON": [],
            "KOMPAS": [],
        }

        grouper = self.subloc.groupby("PAR_ID")
        par_dict = {
            "LOC_ID": [],
            "LOC_NAME": [],
            "X": [],
            "Y": [],
            "ALLE_TYPES": [],
            "START": [],
            "EIND": [],
            "SYSTEEM": [],
            "RAYON": [],
            "KOMPAS": [],
        }

        for loc_id, gdf in grouper:
            caw_code = loc_id[2:-2]
            errors = dict.fromkeys(["LOC_NAME", "GEOMETRY", "SYSTEEM", "RAYON", "KOMPAS"], False)
            fields = dict.fromkeys(par_dict.keys(), None)
            fields["LOC_ID"] = loc_id

            loc_names = np.unique(gdf["LOC_NAME"].str.extract(pat=f"([A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*)").values)

            if len(loc_names) != 1:
                errors["LOC_NAME"] = ",".join(loc_names)
            else:
                fields["LOC_NAME"] = loc_names[0]

            if any([re.match(pattern=loc, string=loc_id) for loc in xy_ignore_df["internalLocation"]]):
                fields["X"], fields["Y"] = next(
                    [row["x"], row["y"]]
                    for index, row in xy_ignore_df.iterrows()
                    if re.match(pattern=row["internalLocation"], string=loc_id)
                )

            else:
                geoms = gdf["geometry"].unique()
                if len(geoms) != 1:
                    errors["GEOMETRY"] = ",".join([f"({geom.x} {geom.y})" for geom in geoms])
                else:
                    fields["X"] = geoms[0].x
                    fields["Y"] = geoms[0].y

            all_types = list(gdf["TYPE"].unique())
            all_types.sort()
            fields["ALLE_TYPES"] = "/".join(all_types)
            fields["START"] = gdf["START"].min()
            fields["EIND"] = gdf["EIND"].max()
            for attribuut in ["SYSTEEM", "RAYON", "KOMPAS"]:
                vals = gdf[attribuut].unique()
                if len(vals) != 1:
                    errors[attribuut] = ",".join(vals)
                else:
                    fields[attribuut] = vals[0]
            if None not in fields.values():
                for key, value in fields.items():
                    par_dict[key].append(value)

            if any(errors.values()):
                hloc_errors["LOC_ID"].append(loc_id)
                hloc_errors["SUB_LOCS"].append(",".join(gdf["LOC_ID"].values))
                for key, value in errors.items():
                    if value is False:
                        value = ""
                    hloc_errors[key].append(value)

        result_df = pd.DataFrame(hloc_errors)
        nr_errors_found = len(result_df)
        if nr_errors_found:
            logger.warning(f"{nr_errors_found} errors in consistency hlocs")
            logger.warning("Hlocs will only be re-written when consistency errors are resolved")
        else:
            logger.info("no consistency errors. Rewrite hlocs from sublocs")
            self._create_hoofdloc_new(par_dict=par_dict)
        return sheet_name, result_df

    def check_expar_errors_intloc_missing(
        self, expar_sheet_name: str = "exPar error", intloc_sheet_name: str = "intLoc missing"
    ) -> Tuple[str, pd.DataFrame, str, pd.DataFrame]:
        """Check on wrong external parameters and missing internal locations. This check returns
        two sheets (name+df), whereas all other checks return one sheet (name+df)."""
        logger.info(f"start {self.check_expar_errors_intloc_missing.__name__}")

        ex_par_errors = {
            "internalLocation": [],
            "locationType": [],
            "exParError": [],
            "types": [],
            "FQ": [],
            "I.X": [],
            "IX.": [],
            "SS./SM.": [],
        }
        int_loc_missing = []
        idmaps = self._get_idmaps(idmap_files=["IdOPVLWATER"])
        idmap_df = pd.DataFrame(idmaps)
        for int_loc, loc_group in idmap_df.groupby("internalLocation"):
            errors = dict.fromkeys(["I.X", "IX.", "FQ", "SS./SM."], False)
            ex_pars = np.unique(loc_group["externalParameter"].values)
            ex_pars_gen = [re.sub(r"\d", ".", ex_par) for ex_par in ex_pars]
            if int_loc in self.hoofdloc["LOC_ID"].values:
                loc_properties = self.hoofdloc[self.hoofdloc["LOC_ID"] == int_loc]
                loc_type = "hoofdloc"
            elif int_loc in self.subloc["LOC_ID"].values:
                loc_properties = self.subloc[self.subloc["LOC_ID"] == int_loc]
                loc_type = "subloc"
                regexes = ["HR.$"]
            elif int_loc in self.waterstandloc["LOC_ID"].values:
                loc_type = "waterstandloc"
            elif int_loc in self.mswloc["LOC_ID"].values:
                loc_type = "mswloc"
            else:
                loc_type = None
                int_loc_missing += [int_loc]
            if loc_type in ["hoofdloc", "subloc"]:
                all_types = loc_properties["ALLE_TYPES"].values[0].split("/")
                all_types = [item.lower() for item in all_types]
            elif loc_type == "waterstandloc":
                all_types = ["waterstandloc"]
            if loc_type == "subloc":
                sub_type = self.subloc[self.subloc["LOC_ID"] == int_loc]["TYPE"].values[0]

                regexes += [
                    j
                    for i in [
                        values for keys, values in constants.EXTERNAL_PARAMETERS_ALLOWED.items() if keys in all_types
                    ]
                    for j in i
                ]

                regexes += list(dict.fromkeys(regexes))
                ex_par_error = [
                    ex_par
                    for ex_par in ex_pars
                    if not any([regex.match(ex_par) for regex in [re.compile(rex) for rex in regexes]])
                ]

                if sub_type == "schuif":
                    if not any([ex_par for ex_par in ex_pars_gen if ex_par in ["SS.", "SM."]]):
                        errors["SS./SM."] = True

                if any([ex_par for ex_par in ex_pars_gen if ex_par in ["I.B", "I.H", "I.L"]]):
                    if not any([ex_par for ex_par in ex_pars_gen if ex_par in ["IB.", "IH.", "IL."]]):
                        errors["IX."] = True

                elif any([ex_par for ex_par in ex_pars_gen if ex_par in ["IB.", "IH.", "IL."]]):
                    errors["I.X"] = True

                if "FQ." in ex_pars_gen:
                    if not any(
                        [ex_par for ex_par in ex_pars_gen if ex_par in ["IB.", "IH.", "IL.", "I.B", "I.H", "I.L"]]
                    ):
                        errors["FQ"] = True

            elif loc_type == "hoofdloc":
                regexes = ["HS.$", "QR.$", "QS.$", "WR", "WS"]
                ex_par_error = [
                    ex_par
                    for ex_par in ex_pars
                    if not any([regex.match(ex_par) for regex in [re.compile(rex) for rex in regexes]])
                ]

            else:
                ex_par_error = []

            if len(ex_par_error) > 0 or any(errors.values()):
                ex_par_errors["internalLocation"].append(int_loc)
                ex_par_errors["locationType"].append(loc_type)
                ex_par_errors["exParError"].append(",".join(ex_par_error))
                ex_par_errors["types"].append(",".join(all_types))
                for key, value in errors.items():
                    ex_par_errors[key].append(value)

        expar_result_df = pd.DataFrame(ex_par_errors)
        intloc_result_df = pd.DataFrame({"internalLocation": int_loc_missing})

        if len(expar_result_df) == 0:
            logger.info("no external parameter errors")
        else:
            logger.warning(f"{len(expar_result_df)} locations with external parameter errors")

        if len(intloc_result_df) == 0:
            logger.info("all internal locations are in locationSets")
        else:
            logger.warning(f"{len(intloc_result_df)} internal locations are not in locationSets")

        return expar_sheet_name, expar_result_df, intloc_sheet_name, intloc_result_df

    def check_expar_missing(self, sheet_name: str = "exPar missing") -> Tuple[str, pd.DataFrame]:
        """Check if external parameters are missing on locations."""
        logger.info(f"start {self.check_expar_missing.__name__}")
        ex_par_missing = {
            "internalLocation": [],
            "exPars": [],
            "QR": [],
            "QS": [],
            "HS": [],
        }
        idmaps = self._get_idmaps(idmap_files=["IdOPVLWATER"])
        idmap_df = pd.DataFrame(data=idmaps)
        for index, row in self.hoofdloc.iterrows():
            missings = dict.fromkeys(["QR", "QS", "HS"], False)
            int_loc = row["LOC_ID"]

            loc_group = next(
                (df for loc, df in idmap_df.groupby("internalLocation") if loc == int_loc), pd.DataFrame(),
            )

            if not loc_group.empty:
                ex_pars = np.unique(ar=loc_group["externalParameter"].values)
                ex_pars_gen = [re.sub(pattern=r"\d", repl=".", string=ex_par) for ex_par in ex_pars]
            else:
                ex_pars = []
                ex_pars_gen = []
            if "HS." not in ex_pars_gen:
                missings["HS"] = True
            if "QR." not in ex_pars_gen:
                missings["QR"] = True
            if "QS." not in ex_pars_gen:
                missings["QS"] = True
            if any(missings.values()):
                ex_par_missing["internalLocation"].append(int_loc)
                ex_par_missing["exPars"].append(",".join(ex_pars))
                for key, value in missings.items():
                    ex_par_missing[key].append(value)

        result_df = pd.DataFrame(ex_par_missing)

        if len(result_df) == 0:
            logger.info("No external parameters missing")
        else:
            logger.warning(f"{len(result_df)} locations with external parameter missing")

        return sheet_name, result_df

    def check_exloc_intloc_consistency(self, sheet_name: str = "exLoc error") -> Tuple[str, pd.DataFrame]:
        """Check if external locations are consistent with internal locations."""
        logger.info(f"start {self.check_exloc_intloc_consistency.__name__}")
        ex_loc_errors = {"internalLocation": [], "externalLocation": []}
        idmaps = self._get_idmaps(idmap_files=["IdOPVLWATER"])
        idmap_df = pd.DataFrame(data=idmaps)
        for loc_group in idmap_df.groupby("externalLocation"):
            int_loc_error = []
            ex_loc = loc_group[0]
            int_locs = np.unique(loc_group[1]["internalLocation"].values)
            if len(ex_loc) == 3:
                if not bool(re.match(pattern="8..$", string=ex_loc)):
                    int_loc_error = [
                        int_loc for int_loc in int_locs if not bool(re.match(pattern=f"...{ex_loc}..$", string=int_loc))
                    ]
                else:
                    for loc_type in ["KW", "OW"]:
                        int_locs_select = [
                            int_loc for int_loc in int_locs if bool(re.match(pattern=f"{loc_type}.", string=int_loc))
                        ]
                        if len(np.unique([int_loc[:-1] for int_loc in int_locs_select])) > 1:
                            int_loc_error += list(int_locs_select)
            if len(ex_loc) == 4:
                if not bool(re.match(pattern=".8..$", string=ex_loc)):
                    int_loc_error += [
                        int_loc for int_loc in int_locs if not bool(re.match(pattern=f"..{ex_loc}..$", string=int_loc))
                    ]
                else:
                    for loc_type in ["KW", "OW"]:
                        int_locs_select = [
                            int_loc for int_loc in int_locs if bool(re.match(pattern=f"{loc_type}.", string=int_loc))
                        ]
                        if len(np.unique([int_loc[:-1] for int_loc in int_locs_select])) > 1:
                            int_loc_error += list(int_locs_select)
            # renier
            if "exLoc_ignore" in self.consistency.keys():
                if int(ex_loc) in self.consistency["exLoc_ignore"]["externalLocation"].values:
                    int_loc_error = [
                        int_loc
                        for int_loc in int_loc_error
                        if int_loc
                        not in self.consistency["exLoc_ignore"][
                            self.consistency["exLoc_ignore"]["externalLocation"] == int(ex_loc)
                        ]["internalLocation"].values
                    ]

            for int_loc in int_loc_error:
                ex_loc_errors["internalLocation"].append(int_loc)
                ex_loc_errors["externalLocation"].append(ex_loc)

        result_df = pd.DataFrame(ex_loc_errors)

        if len(result_df) == 0:
            logger.info("all external and internal locations consistent")
        else:
            logger.warning(f"{len(result_df)} external locations inconsistent with internal locations")

        return sheet_name, result_df

    def check_timeseries_logic(self, sheet_name: str = "timeSeries error") -> Tuple[str, pd.DataFrame]:
        """Check if timeseries are consistent with internal locations and parameters."""
        logger.info(f"start {self.check_timeseries_logic.__name__}")
        if "TS800_ignore" in self.consistency.keys():
            ts_ignore_df = self.consistency["TS800_ignore"]
        else:
            ts_ignore_df = pd.DataFrame({"internalLocation": [], "externalLocation": []})

        idmaps = self._get_idmaps(idmap_files=["IdOPVLWATER"])
        idmap_df = pd.DataFrame(data=idmaps)

        idmap_subloc_df = idmap_df[idmap_df["internalLocation"].isin(values=self.subloc["LOC_ID"].values)]

        idmap_subloc_df.loc[:, "type"] = idmap_subloc_df["internalLocation"].apply(
            func=(lambda x: self.subloc[self.subloc["LOC_ID"] == x]["TYPE"].values[0])
        )

        idmap_subloc_df.loc[:, "loc_groep"] = idmap_subloc_df["internalLocation"].apply(func=(lambda x: x[0:-1]))

        ts_errors = {
            "internalLocation": [],
            "eind": [],
            "internalParameters": [],
            "externalParameters": [],
            "externalLocations": [],
            "type": [],
            "fout": [],
        }

        for loc_group, group_df in idmap_subloc_df.groupby("loc_groep"):
            ex_locs = np.unique(group_df["externalLocation"].values)
            ex_locs_dict = {ex_loc: idx for idx, ex_loc in enumerate(ex_locs)}
            split_ts = [
                key
                for key in ex_locs_dict.keys()
                if any([regex.match(string=key) for regex in [re.compile(pattern=rex) for rex in ["8..", ".8.."]]])
            ]

            ex_locs_skip = ts_ignore_df[ts_ignore_df["internalLocation"].isin(values=group_df["internalLocation"])][
                "externalLocation"
            ]

            split_ts = [key for key in split_ts if str(key) not in ex_locs_skip.values.astype(np.str)]

            # TODO: whuuuuaaatt happens here!? :)
            # TODO: also use .get()
            ex_locs_dict = {
                k: (ex_locs_dict[k[1:]] if (k[1:] in ex_locs_dict.keys()) and (k not in split_ts) else v)
                for (k, v) in ex_locs_dict.items()
            }

            org_uniques = np.unique([val for key, val in ex_locs_dict.items() if key not in split_ts])
            if len(org_uniques) == 1 and len(split_ts) == 1:
                ex_locs_dict = {k: (org_uniques[0] if k in split_ts else v) for (k, v) in ex_locs_dict.items()}

            group_df["ex_loc_group"] = group_df["externalLocation"].apply(func=(lambda x: ex_locs_dict[x]))

            for int_loc, loc_df in group_df.groupby("internalLocation"):
                sub_type = self.subloc[self.subloc["LOC_ID"] == int_loc]["TYPE"].values[0]

                date_str = self.subloc[self.subloc["LOC_ID"] == int_loc]["EIND"].values[0]

                end_time = None
                try:
                    end_time = pd.to_datetime(date_str)
                except pd.errors.OutOfBoundsDatetime as err:
                    # TODO: @renier: handle out of bounds in case date_str is e.g. '32000101'
                    logger.error(err)

                ex_pars = np.unique(loc_df["externalParameter"].values)
                int_pars = np.unique(loc_df["internalParameter"].values)
                ex_locs = np.unique(loc_df["externalLocation"].values)
                if sub_type in ["krooshek", "debietmeter"]:
                    if any([re.match(pattern="HR.", string=ex_par) for ex_par in ex_pars]):
                        ts_errors["internalLocation"].append(int_loc)
                        ts_errors["eind"].append(end_time)
                        ts_errors["internalParameters"].append(",".join(int_pars))
                        ts_errors["externalParameters"].append(",".join(ex_pars))
                        ts_errors["externalLocations"].append(",".join(ex_locs))
                        ts_errors["type"].append(sub_type)
                        ts_errors["fout"].append(f"{sub_type} met stuurpeil")

                else:
                    if not any([re.match(pattern="HR.", string=ex_par) for ex_par in ex_pars]):
                        if any(
                            [
                                re.match(pattern="HR.", string=ex_par)
                                for ex_par in np.unique(group_df["externalParameter"])
                            ]
                        ):
                            if sub_type not in ["totaal", "vispassage"]:
                                if pd.Timestamp.now() < end_time:
                                    sp_locs = np.unique(
                                        group_df[group_df["externalParameter"].str.match("HR.")]["internalLocation"]
                                    )
                                    ts_errors["internalLocation"].append(int_loc)
                                    ts_errors["eind"].append(end_time)
                                    ts_errors["internalParameters"].append(",".join(int_pars))
                                    ts_errors["externalParameters"].append(",".join(ex_pars))
                                    ts_errors["externalLocations"].append(",".join(ex_locs))
                                    ts_errors["type"].append(sub_type)
                                    ts_errors["fout"].append(f"{sub_type} zonder stuurpeil {','.join(sp_locs)} wel")
                    else:
                        time_series = loc_df.groupby(["ex_loc_group", "externalParameter"])
                        sp_series = [
                            series for series in time_series if bool(re.match(pattern="HR.", string=series[0][1]))
                        ]
                        for idx, series in enumerate(sp_series):
                            ex_par = series[0][1]
                            ex_locs = series[1]["externalLocation"]
                            int_par = np.unique(series[1]["internalParameter"])
                            if len(int_par) > 1:
                                ts_errors["internalLocation"].append(int_loc)
                                ts_errors["eind"].append(end_time)
                                ts_errors["internalParameters"].append(",".join(int_pars))

                                ts_errors["externalParameters"].append(",".join(ex_pars))

                                ts_errors["externalLocations"].append(",".join(ex_locs))
                                ts_errors["type"].append(sub_type)
                                ts_errors["fout"].append(
                                    (
                                        f'{",".join(int_par)} coupled to 1 sp-series ('
                                        f"exPar: {ex_par}, exLoc(s)): "
                                        f'{",".join(ex_locs)}'
                                    )
                                )
                            other_series = [series for idy, series in enumerate(sp_series) if idy != idx]

                            other_int_pars = [np.unique(series[1]["internalParameter"]) for series in other_series]

                            if len(other_int_pars) > 0:
                                other_int_pars = np.concatenate(other_int_pars)

                            conflicting_pars = [par for par in int_par if par in other_int_pars]

                            if len(conflicting_pars) > 0:
                                # TODO: @daniel comment kan weg?
                                # 2 sp series coupled to the same fews parameter
                                ts_errors["internalLocation"].append(int_loc)
                                ts_errors["eind"].append(end_time)
                                ts_errors["internalParameters"].append(",".join(int_pars))

                                ts_errors["externalParameters"].append(",".join(ex_pars))

                                ts_errors["externalLocations"].append(",".join(ex_locs))
                                ts_errors["type"].append(sub_type)
                                ts_errors["fout"].append(
                                    (
                                        f'{",".join(conflicting_pars)} coupled to  '
                                        f"sp-serie (exPar: {ex_par}, exLoc(s)):"
                                        f'{",".join(ex_locs)}'
                                    )
                                )

        result_df = pd.DataFrame(ts_errors)
        if len(result_df) == 0:
            logger.info("logical coupling of all timeseries to internal locations/parameters")
        else:
            logger.warning(f"{len(result_df)} timeseries coupled illogical to internal locations/parameters")

        return sheet_name, result_df

    def check_validation_rules(self, sheet_name: str = "validation error") -> Tuple[str, pd.DataFrame]:
        """Check if validation rules are consistent."""
        logger.info(f"start {self.check_validation_rules.__name__}")
        valid_errors = {
            "internalLocation": [],
            "start": [],
            "eind": [],
            "internalParameters": [],
            "fout_type": [],
            "fout_beschrijving": [],
        }

        locs_mapping = {
            "hoofdlocaties": "hoofdloc",
            "sublocaties": "subloc",
            "waterstandlocaties": "waterstandloc",
            "mswlocaties": "mswloc",
        }

        locations_dict = xml_to_dict(xml_filepath=self.fews_config.RegionConfigFiles["LocationSets"])
        location_sets = locations_dict["locationSets"]["locationSet"]

        for set_name in constants.VALIDATION_RULES.keys():
            location_set_meta = next(
                loc_set for loc_set in location_sets if loc_set["id"] == self.location_sets[set_name]["id"]
            )["csvFile"]

            location_set_gdf = getattr(self, locs_mapping[set_name])
            attrib_files = location_set_meta["attributeFile"]
            if not isinstance(attrib_files, list):
                attrib_files = [attrib_files]

            attrib_files = [attrib_file for attrib_file in attrib_files if "attribute" in attrib_file.keys()]

            for attrib_file in attrib_files:
                attribs = attrib_file["attribute"]
                join_id = attrib_file["id"].replace("%", "")
                if not isinstance(attrib_file["attribute"], list):
                    attribs = [attribs]
                attribs = [attrib["number"].replace("%", "") for attrib in attribs if "number" in attrib.keys()]

                attrib_df = pd.read_csv(
                    self.fews_config.MapLayerFiles[attrib_file["csvFile"].replace(".csv", "")],
                    sep=None,
                    engine="python",
                )

                attrib_df.rename(columns={join_id: "LOC_ID"}, inplace=True)
                drop_cols = [col for col in attrib_df if col not in attribs + ["LOC_ID"]]

                attrib_df = attrib_df.drop(columns=drop_cols, axis=1)
                location_set_gdf = location_set_gdf.merge(attrib_df, on="LOC_ID", how="outer")

            validation_rules = constants.VALIDATION_RULES[set_name]
            validaton_attributes = get_validation_attribs(validation_rules)

            idmaps = self._get_idmaps(idmap_files=["IdOPVLWATER"])
            idmap_df = pd.DataFrame(data=idmaps)

            params_df = pd.DataFrame.from_dict(
                data={
                    int_loc: [df["internalParameter"].values] for int_loc, df in idmap_df.groupby("internalLocation")
                },
                orient="index",
                columns=["internalParameters"],
            )

            for (idx, row) in location_set_gdf.iterrows():
                int_loc = row["LOC_ID"]
                row = row.dropna()
                if int_loc in params_df.index:
                    int_pars = np.unique(params_df.loc[int_loc]["internalParameters"])
                else:
                    int_pars = []

                attribs_required = get_validation_attribs(validation_rules, int_pars)
                attribs_missing = [attrib for attrib in attribs_required if attrib not in row.keys()]

                attribs_obsolete = [
                    attrib
                    for attrib in validaton_attributes
                    if (attrib not in attribs_required) and (attrib in row.keys())
                ]

                for key, value in {"missend": attribs_missing, "overbodig": attribs_obsolete}.items():
                    if len(value) > 0:
                        valid_errors["internalLocation"] += [int_loc]
                        valid_errors["start"] += [row["START"]]
                        valid_errors["eind"] += [row["EIND"]]
                        valid_errors["internalParameters"] += [",".join(int_pars)]
                        valid_errors["fout_type"] += [key]
                        valid_errors["fout_beschrijving"] += [",".join(value)]
                for validation_rule in validation_rules:
                    errors = {"fout_type": None, "fout_beschrijving": []}
                    param = validation_rule["parameter"]
                    if any(re.match(pattern=param, string=int_par) for int_par in int_pars):
                        rule = validation_rule["extreme_values"]
                        rule = sort_validation_attribs(rule)
                        if all(key in ["hmax", "hmin"] for key in rule.keys()):
                            for hmin, hmax in zip(rule["hmin"], rule["hmax"]):
                                if all(attrib in row.keys() for attrib in [hmin, hmax]):
                                    if row[hmax] < row[hmin]:
                                        errors["fout_type"] = "waarde"
                                        errors["fout_beschrijving"] += [f"{hmax} < {hmin}"]

                        elif all(key in rule.keys() for key in ["hmax", "smax", "smin", "hmin"]):
                            hmax = rule["hmax"][0]
                            hmin = rule["hmin"][0]
                            for smin, smax in zip(rule["smin"], rule["smax"]):
                                if all(attrib in row.keys() for attrib in [smin, smax]):
                                    if row[smax] <= row[smin]:
                                        errors["fout_type"] = "waarde"
                                        errors["fout_beschrijving"] += [f"{smax} <= {smin}"]

                                    if row[hmax] < row[smax]:
                                        errors["fout_type"] = "waarde"
                                        errors["fout_beschrijving"] += [f"{'hmax'} < {smax}"]

                                    if row[smin] < row[hmin]:
                                        errors["fout_type"] = "waarde"
                                        errors["fout_beschrijving"] += [f"{smin} < {hmin}"]

                    valid_errors["internalLocation"] += [row["LOC_ID"]] * len(errors["fout_beschrijving"])

                    valid_errors["start"] += [row["START"]] * len(errors["fout_beschrijving"])

                    valid_errors["eind"] += [row["EIND"]] * len(errors["fout_beschrijving"])

                    valid_errors["internalParameters"] += [",".join(int_pars)] * len(errors["fout_beschrijving"])

                    valid_errors["fout_type"] += [errors["fout_type"]] * len(errors["fout_beschrijving"])

                    valid_errors["fout_beschrijving"] += errors["fout_beschrijving"]

        # TODO: @renier: verify that
        #  self.consistency[sheet_name] = pd.DataFrame(valid_errors)
        #  self.consistency[sheet_name] = self.consistency[sheet_name].drop_duplicates()
        # is the same as:
        result_df = pd.DataFrame(valid_errors).drop_duplicates()

        if len(result_df) == 0:
            logger.info("no missing incorrect validation rules")
        else:
            logger.warning(f"{len(result_df)} validation rules contain errors/are missing")
        return sheet_name, result_df

    def check_intpar_expar_consistency(self, sheet_name: str = "par mismatch") -> Tuple[str, pd.DataFrame]:
        """Check if internal and external parameters are consistent."""
        logger.info(f"start {self.check_intpar_expar_consistency.__name__}")
        par_errors = {
            "internalLocation": [],
            "internalParameter": [],
            "externalParameter": [],
            "fout": [],
        }

        idmaps = self._get_idmaps(idmap_files=["IdOPVLWATER"])
        idmap_df = pd.DataFrame(data=idmaps)
        for idx, row in idmap_df.iterrows():
            error = None
            ext_par = [
                mapping["external"]
                for mapping in constants.PARAMETER_MAPPING
                if re.match(pattern=f'{mapping["internal"]}[0-9]', string=row["internalParameter"])
            ]

            if ext_par:
                if not any(re.match(pattern=par, string=row["externalParameter"]) for par in ext_par):
                    error = "parameter mismatch"
            else:
                error = "pars not included in config"
            if error:
                par_errors["internalLocation"].append(row["internalLocation"])
                par_errors["internalParameter"].append(row["internalParameter"])
                par_errors["externalParameter"].append(row["externalParameter"])
                par_errors["fout"].append(error)

        result_df = pd.DataFrame(par_errors)

        if len(result_df) == 0:
            logger.info("no regex errors for internal and external parameters")
        else:
            logger.warning(f"{len(result_df)} regex errors for internal and external parameters")
        return sheet_name, result_df

    def check_location_set_errors(self, sheet_name: str = "locSet error") -> Tuple[str, pd.DataFrame]:
        """Check on errors in locationsets."""
        logger.info(f"start {self.check_location_set_errors.__name__}")
        xy_ignore_df = self.consistency["xy_ignore"]
        location_sets = self.location_sets

        loc_set_errors = {
            "locationId": [],
            "caw_code": [],
            "caw_name": [],
            "csv_file": [],
            "location_name": [],
            "type": [],
            "functie": [],
            "name_error": [],
            "caw_name_inconsistent": [],
            "missing_in_map": [],
            "missing_in_set": [],
            "missing_peilschaal": [],
            "missing_hbov": [],
            "missing_hben": [],
            "missing_hbovps": [],
            "missing_hbenps": [],
            "missing_hloc": [],
            "xy_not_same": [],
        }

        sets = {
            "waterstandlocaties": "WATERSTANDLOCATIES",
            "sublocaties": "KUNSTWERKEN",
            "hoofdlocaties": "KUNSTWERKEN",
        }

        for set_name, section_name in sets.items():
            logger.info(set_name)
            location_set = location_sets[set_name]
            location_gdf = location_set["gdf"]
            csv_file = self.fews_config.location_sets[location_set["id"]]["csvFile"]["file"]
            int_locs = []

            for idmap in ["IdOPVLWATER", "IdOPVLWATER_HYMOS"]:
                for section in constants.IDMAP_SECTIONS[idmap][section_name]:
                    int_locs += [
                        item["internalLocation"]
                        for item in xml_to_dict(self.fews_config.IdMapFiles[idmap], **section)["idMap"]["map"]
                    ]

            if set_name == "sublocaties":
                int_locs = [loc for loc in int_locs if loc[-1] != "0"]
                par_gdf = location_sets["hoofdlocaties"]["gdf"]

            elif set_name == "hoofdlocaties":
                int_locs = [loc for loc in int_locs if loc[-1] == "0"]

            for idx, row in list(location_gdf.iterrows()):
                error = {
                    "name_error": False,
                    "caw_name_inconsistent": False,
                    "missing_in_map": False,
                    "type": "",
                    "functie": "",
                    "missing_in_set": False,
                    "missing_peilschaal": False,
                    "missing_hbov": False,
                    "missing_hben": False,
                    "missing_hbovps": False,
                    "missing_hbenps": False,
                    "missing_hloc": False,
                    "xy_not_same": False,
                }

                loc_id = row["LOC_ID"]
                caw_code = loc_id[2:-2]
                loc_name = row["LOC_NAME"]
                caw_name = ""

                if set_name == "sublocaties":

                    loc_functie = row["FUNCTIE"]
                    sub_type = row["TYPE"]

                    if sub_type in [
                        "afsluiter",
                        "debietmeter",
                        "krooshek",
                        "vispassage",
                    ]:
                        if not re.match(pattern=f"[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*-{sub_type}", string=loc_name):
                            error["name_error"] = True

                    else:
                        if not re.match(
                            pattern=f"[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*-{sub_type}[0-9]_{loc_functie}",
                            string=loc_name,
                        ):
                            error["name_error"] = True

                    if not error["name_error"]:
                        caw_name = re.match(pattern="([A-Z0-9 ]*)_", string=loc_name).group(1)
                        if not all(
                            location_gdf[location_gdf["LOC_ID"].str.match(f"..{caw_code}")]["LOC_NAME"].str.match(
                                f"({caw_name}_{caw_code}-K)"
                            )
                        ):
                            error["caw_name_inconsistent"] = True

                    if row["HBOV"] not in location_sets["waterstandlocaties"]["gdf"]["LOC_ID"].values:
                        error["missing_hbov"] = True

                    if row["HBEN"] not in location_sets["waterstandlocaties"]["gdf"]["LOC_ID"].values:
                        error["missing_hben"] = True

                    if row["HBOVPS"] not in location_sets["peilschalen"]["gdf"]["LOC_ID"].values:
                        error["missing_hbovps"] = True

                    if row["HBENPS"] not in location_sets["peilschalen"]["gdf"]["LOC_ID"].values:
                        error["missing_hbenps"] = True

                    if row["PAR_ID"] not in location_sets["hoofdlocaties"]["gdf"]["LOC_ID"].values:
                        error["missing_hloc"] = True

                    else:
                        if not any([re.match(pattern=loc, string=loc_id) for loc in xy_ignore_df["internalLocation"]]):
                            if (
                                not par_gdf[par_gdf["LOC_ID"] == row["PAR_ID"]]["geometry"]
                                .values[0]
                                .equals(row["geometry"])
                            ):
                                error["xy_not_same"] = True

                    if any(error.values()):
                        error["type"] = sub_type
                        error["functie"] = loc_functie

                elif set_name == "hoofdlocaties":
                    if not re.match(pattern=f"[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*", string=loc_name):
                        error["name_error"] = True

                elif set_name == "waterstandlocaties":
                    if not re.match(pattern=f"[A-Z0-9 ]*_{caw_code}-w_.*", string=loc_name):
                        error["name_error"] = True

                    if not error["name_error"]:
                        caw_name = re.match(pattern="([A-Z0-9 ]*)_", string=loc_name).group(1)
                        if not all(
                            location_gdf[location_gdf["LOC_ID"].str.match(f"..{caw_code}")]["LOC_NAME"].str.match(
                                f"({caw_name}_{caw_code}-w)"
                            )
                        ):
                            error["caw_name_inconsistent"] = True

                    if row["PEILSCHAAL"] not in location_sets["peilschalen"]["gdf"]["LOC_ID"].values:
                        error["missing_peilschaal"] = True

                    if loc_id not in int_locs:
                        error["missing_in_map"] = True

                if any(error.values()):
                    loc_set_errors["locationId"].append(loc_id)
                    loc_set_errors["caw_name"].append(caw_name)
                    loc_set_errors["caw_code"].append(caw_code)
                    loc_set_errors["csv_file"].append(csv_file)
                    loc_set_errors["location_name"].append(loc_name)
                    for key, value in error.items():
                        loc_set_errors[key].append(value)

        result_df = pd.DataFrame(loc_set_errors)
        if len(result_df) == 0:
            logger.info("no errors in locationSets")
        else:
            logger.warning(f"{len(result_df)} errors in locationSets")
        return sheet_name, result_df

    def write_excel(self) -> None:
        """Write consistency to excel.
        1. copy file consistency_input_xlsx to consistency_output_xlsx
        2. empty all warning sheets of consistency_output_xlsx
        3. save results (incl summary) to consistency_output_xlsx
        """
        consistency_output_xlsx_path = constants.PathConstants.output_dir.path / "consistency_output.xlsx"
        shutil.copy(src=constants.PathConstants.consistency_input_xlsx.path, dst=consistency_output_xlsx_path)

        # remove all xlsx sheets that will be filled with new data
        book = load_workbook(filename=consistency_output_xlsx_path)
        for worksheet in book.worksheets:
            if worksheet.title not in constants.FIXED_SHEETS:
                book.remove(worksheet)

        # create summary dict
        inhoudsopgave = self.consistency["inhoudsopgave"]
        inhoudsopgave.index = inhoudsopgave["werkblad"]
        summary = {
            key: len(df)
            for key, df in self.consistency.items()
            if key not in constants.FIXED_SHEETS + ["inhoudsopgave"]
        }

        # TODO: @renier: remove this integration test
        from mptconfig.tmp import validate_expected_summary

        validate_expected_summary(new_summary=summary)

        # add summary
        worksheet = book.create_sheet(title="samenvatting", index=1)
        worksheet.sheet_properties.tabColor = "92D050"
        worksheet.append(["controle", "aantal", "beschrijving"])
        max_row = f"{worksheet.max_row}"
        for cell in worksheet[max_row]:
            cell.font = Font(bold=True)

        for key, value in summary.items():
            beschrijving = inhoudsopgave.loc[key]["beschrijving"]
            worksheet.append([key, value, beschrijving])
            if value > 0 and key != "mpt":
                worksheet[worksheet.max_row][1].fill = PatternFill(fgColor="FF0000", fill_type="solid")
            else:
                worksheet[worksheet.max_row][1].fill = PatternFill(fgColor="92D050", fill_type="solid")

        worksheet.column_dimensions["A"].width = 40
        worksheet.column_dimensions["C"].width = 100
        worksheet.auto_filter.ref = worksheet.dimensions

        # add warning sheets
        xls_writer = pd.ExcelWriter(path=consistency_output_xlsx_path, engine="openpyxl")
        xls_writer.book = book

        for sheet_name in summary.keys():
            df = self.consistency[sheet_name]
            if df.empty:
                continue
            if df.index.name is None:
                df.to_excel(xls_writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(xls_writer, sheet_name=sheet_name, index=True)
            worksheet = xls_writer.sheets[sheet_name]
            for col in worksheet.columns:
                worksheet.column_dimensions[col[0].column_letter].width = 20
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.freeze_panes = worksheet["B2"]
            if not df.empty:
                if sheet_name == "mpt":
                    worksheet.sheet_properties.tabColor = "92D050"
                else:
                    worksheet.sheet_properties.tabColor = "FF0000"

        xls_writer.book.active = xls_writer.book["samenvatting"]
        xls_writer.save()

    def write_csvs(self) -> None:
        """Write locationSets to csv files."""
        assert not self.consistency["mpt"].empty

        date_threshold = self.consistency["mpt"]["ENDDATE"].max() - pd.Timedelta(weeks=26)

        location_sets = {
            key: value
            for key, value in self.location_sets.items()
            if value["id"] in ["OPVLWATER_HOOFDLOC", "OPVLWATER_WATERSTANDEN_AUTO", "OPVLWATER_SUBLOC"]
        }

        for key, value in location_sets.items():
            logger.info(f"writing CSV for set: {key}. This may take a while")
            gdf = value["gdf"]
            df = gdf.drop("geometry", axis=1)
            df[["START", "EIND"]] = df.apply(
                func=update_date, args=(self.consistency["mpt"], date_threshold), axis=1, result_type="expand"
            )

            if value["id"] == "OPVLWATER_WATERSTANDEN_AUTO":
                grouper = self.mpt_histtags.groupby(["fews_locid"])
                # leave it HIST_TAG (instead of HISTTAG), as that is what OPVLWATER_WATERSTANDEN_AUTO.csv expects..
                df["HIST_TAG"] = df.apply(func=update_histtag, args=[grouper], axis=1, result_type="expand")

            elif value["id"] == "OPVLWATER_SUBLOC":
                grouper = df.groupby(["PAR_ID"])
                par_types_df = grouper["TYPE"].unique().apply(func=lambda x: sorted(x)).transform(lambda x: "/".join(x))

                df["PAR_ID"] = gdf["LOC_ID"].str[0:-1] + "0"
                df["ALLE_TYPES"] = df["PAR_ID"].apply(func=lambda x: par_types_df.loc[x])
                df[["HBOVPS", "HBENPS"]] = df.apply(func=self._update_staff_gauge, axis=1, result_type="expand")

            csv_file_name = self.fews_config.location_sets[value["id"]]["csvFile"]["file"]
            csv_file_path = constants.PathConstants.output_dir.path / csv_file_name
            if not csv_file_path.suffix:
                csv_file_path = Path(f"{csv_file_path}.csv")
            df.to_csv(csv_file_path, index=False)

    def run(self):
        sheet_name, sheet_df = self.check_idmap_sections()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_ignored_histtags()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_missing_histtags()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_double_idmaps()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.histtags_to_mpt()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_missing_pars()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_hloc_consistency()
        self.consistency[sheet_name] = sheet_df

        sheet_name1, sheet_df1, sheet_name2, sheet_df2 = self.check_expar_errors_intloc_missing()
        self.consistency[sheet_name1] = sheet_df1
        self.consistency[sheet_name2] = sheet_df2

        sheet_name, sheet_df = self.check_expar_missing()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_exloc_intloc_consistency()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_timeseries_logic()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_validation_rules()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_intpar_expar_consistency()
        self.consistency[sheet_name] = sheet_df

        sheet_name, sheet_df = self.check_location_set_errors()
        self.consistency[sheet_name] = sheet_df

        # write excel file
        self.write_excel()

        # write csv files
        self.write_csvs()
