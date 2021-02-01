# -*- coding: utf-8 -*-
__title__ = 'histTags2mpt'
__description__ = 'to evaluate a HDSR FEWS-config with a csv with CAW histTags'
__version__ = '0.1'
__author__ = 'Daniel Tollenaar'
__author_email__ = 'daniel@d2hydro.nl'
__license__ = 'MIT License'

'''
ToDo:
    - instellingen verplaatsen naar config.ini
    - logging ook in bestand opslaan
'''

import configparser
from fews_utilities import Config, xml_to_dict
from pathlib import Path
import json
import numpy as np
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import sys
import shutil
import re
from collections.abc import Iterable
from shapely.geometry import Point

pd.options.mode.chained_assignment = None

#%% instellingen
# layout excel spreadsheet
fixed_sheets = ['histTag_ignore',
                'inhoudsopgave',
                'exLoc_ignore',
                'TS800_ignore',
                'xy_ignore']

warning_sheets = ['histTags_noMatch',
                  'histTags_ignore_match',
                  'dubbele idmaps',
                  'idmap v sectie',
                  'exPar error',
                  'exPar missing',
                  'intLoc missing',
                  'exLoc error',
                  'timeSeries error',
                  'validation error',
                  'par mismatch',
                  'locSet error',
                  'hloc error']

idmap_files = ['IdOPVLWATER',
              'IdOPVLWATER_HYMOS',
              'IdHDSR_NSC',
              'IdOPVLWATER_WQ',
              'IdGrondwaterCAW']

# secties in idmap files
idmap_sections = {'IdOPVLWATER':{'KUNSTWERKEN':[{'section_start': '<!--KUNSTWERK SUBLOCS (old CAW id)-->',
                                                 'section_end': '<!--WATERSTANDSLOCATIES (old CAW id)-->'},
                                                {'section_start': '<!--KUNSTWERK SUBLOCS (new CAW id)-->',
                                                 'section_end':'<!--WATERSTANDSLOCATIES (new CAW id)-->'}],
                          'WATERSTANDLOCATIES':[{'section_start': '<!--WATERSTANDSLOCATIES (old CAW id)-->',
                                                 'section_end': '<!--MSW (old CAW id)-->'},
                                                {'section_start': '<!--WATERSTANDSLOCATIES (new CAW id)-->',
                                                'section_end': '<!--MSW (new CAW id)-->'}],
                          'MSWLOCATIES':[{'section_start': '<!--MSW (new CAW id)-->'}]},
                  'IdOPVLWATER_HYMOS':{'KUNSTWERKEN':[{'section_end':'<!--WATERSTANDSLOCATIES-->'}],
                                       'WATERSTANDLOCATIES':[{'section_start': '<!--WATERSTANDSLOCATIES-->',
                                                              'section_end':'<!--OVERIG-->'}]}
                  }

# exParameters per sub-loc type

expars_allowed = {'pompvijzel': ['FQ.$', 'I.B$', 'IB.$', 'I.H$', 'IH.$', 'I.L$', 'IL.$', 'Q.$' , 'TT.$'],
                  'stuw': ['SW.$', 'Q.$', 'ES.$'],
                  'schuif': ['ES.$', 'SP.$', 'SS.$', 'Q.$', 'SM.$'],
                  'afsluiter': ['ES.$'],
                  'debietmeter': ['Q.$'],
                  'vispassage': ['ES.$', 'SP.$', 'SS.$', 'Q.$'],
                  'krooshek': ['HB.$', 'HO.$'],
                  'waterstand': ['HB.$', 'HO.$', 'H$']}

#%% functies
def idmap2tags(row,idmap):
    '''functie voor het toevoegen van fews-locatie-ids aan de hist_tags data-frame in de apply-method'''
     
    exloc, expar = row['serie'].split('_',1)
    fews_locs = [col['internalLocation'] 
                   for col in idmap
                   if col['externalLocation'] == exloc 
                   and col['externalParameter'] == expar]
    
    if len(fews_locs) == 0:
        fews_locs = np.NaN   

    return fews_locs

def update_hlocs(row):
    '''functie voor het toevoegen van start en end-date op data-frame van hoofdloc in de apply-method'''
    
    loc_id = row.name
    start_date = row['STARTDATE']
    end_date = row['ENDDATE']
    
    if loc_id in h_locs:
        start_date = mpt_df[mpt_df.index.str.contains(loc_id[0:-1])]['STARTDATE'].dropna().min()
        end_date = mpt_df[mpt_df.index.str.contains(loc_id[0:-1])]['ENDDATE'].dropna().max()
    
    return start_date, end_date 

def flatten(l):
    '''functie voor het platslaan van een onregemlatige iterable van lijsten'''
    for el in l:
        if isinstance(el, Iterable) and not isinstance(el, (str, bytes)):
            yield from flatten(el)
        else:
            yield el
            
            
def get_attribs(validation_rules,int_pars=None,loc_type=None):
    '''functie voor het ophalen van attributen uit validation_rules'''
    if int_pars is None:
        int_pars = [rule['parameter'] for rule in validation_rules]
    result = []
    for rule in validation_rules:
        if 'type' in rule.keys():
            if rule['type'] == loc_type:
                if any(re.match(rule['parameter'],int_par) for int_par in int_pars):
                    for key,attribute in rule['extreme_values'].items():
                        if isinstance(attribute,list):
                            result += [value['attribute'] for  value in attribute]
                        else:
                            result += [attribute]
        elif any(re.match(rule['parameter'],int_par) for int_par in int_pars):
            for key,attribute in rule['extreme_values'].items():
                if isinstance(attribute,list):
                    result += [value['attribute'] for  value in attribute]
                else:
                    result += [attribute]
    return result

#%% initialisatie
workdir = Path(__file__).parent
logging.basicConfig(level=os.environ.get("LOGLEVEL", "INFO"))

#inlezen paden vanuit inifile
config_json = Path(r'..\config\config.json')
if config_json.exists():
    with open(config_json) as src:
        config = json.load(src)
else:
    logging.error(f'{config_json} does not exist')
    sys.exit()

#controleren of paden bestaan
for key, path in config['paden'].items():
    path = Path(path)
    if not path.is_absolute():
        path = workdir.joinpath(path).resolve()
    if path.exists():
        config['paden'][key] = path
    else:
        if path.suffix == '':
            logging.warning(f'{path} bestaat niet, map wordt aangemaakt')
            path.mkdir()
        else:
            logging.error(f'{path} bestaat niet. Specificeer het juiste path in config.ini')
            sys.exit()

locals().update(config['paden'])
consistency_out_xlsx = consistency_xlsx.parent.joinpath(f'{consistency_xlsx.stem}_uit.xlsx')

#%% inlezen config-excel
# kopieeren van consistency workbook naar output
try:
    shutil.copyfile(consistency_xlsx, consistency_out_xlsx)
except Exception as e: 
    logging.error(e) 
    sys.exit()

consistency_df = pd.read_excel(consistency_xlsx,sheet_name=None,engine='openpyxl')
if not (('histTag_ignore' in consistency_df.keys()) | (mpt_ignore_csv != None)):
    logging.error(f'specificeer een histTag_ignore werkblad in {consistency_xlsx} of een csv-file in {config_json}')
    sys.exit()
    
# weggooien van alle output-sheets, behalve degenen opgeschoond
consistency_df = {key:value for key,value in consistency_df.items() if key in fixed_sheets}

#%% inlezen idmap-files
fews_config = Config(fews_config)
idmap_dict = {idmap:xml_to_dict(fews_config.IdMapFiles[idmap])['idMap']['map'] 
           for idmap in idmap_files}
idmap_total = [j for i in idmap_dict.values() for j in i]

#%% inlezen locationSets locationSets

location_sets = {location_set:{'id':config['location_sets'][location_set],
                               'gdf':fews_config.get_locations(config['location_sets'][location_set])} 
                 for location_set in config['location_sets']}
   
#%% controle op KW/OW
logging.info('controle op KW/OW locaties in juiste sectie')
consistency_df['idmap v sectie'] = pd.DataFrame(columns=['bestand',
                                                    'externalLocation',
                                                    'externalParameter',
                                                    'internalLocation',
                                                    'internalParameter',
                                                    ])
idmap = 'IdOPVLWATER'
idmap_subsecs = idmap_sections[idmap]
for section_type, sections in idmap_subsecs.items():
    for section in sections:
        if section_type == 'KUNSTWERKEN':
            prefix = 'KW'
        if section_type == 'WATERSTANDLOCATIES':
            prefix = 'OW'
        if section_type == 'MSWLOCATIES':
            prefix = '(OW|KW)'
        pattern = f'{prefix}\d{{6}}$'
        idmap_wrong_section = [idmap for idmap in xml_to_dict(fews_config.IdMapFiles[idmap],**section)['idMap']['map'] 
                               if not bool(re.match(pattern,idmap['internalLocation']))]
        if len(idmap_wrong_section):
            section_start = section['section_start'] if 'section_start' in section.keys() else ''
            section_end = section['section_end'] if 'section_end' in section.keys() else ''
            logging.warning('{} internalLocations anders dan {}XXXXXX tussen {} en {} in {}'.format(len(idmap_wrong_section),
                                                                                                   prefix,
                                                                                                   section_start,
                                                                                                   section_end,
                                                                                                   idmap))
            df = pd.DataFrame(idmap_wrong_section)
            df['sectie'] = section_start
            df['bestand'] = idmap
            consistency_df['idmap v sectie'] = pd.concat([consistency_df['idmap v sectie'], df], axis=0)

#%% inlezen hist tags & ignore lijst
logging.info('zoeken naar missende histTags in idmaps')
dtype_cols = ['total_min_start_dt', 'total_max_end_dt']
hist_tags_org_df = pd.read_csv(hist_tags_csv,
                           parse_dates = dtype_cols,
                           sep = ';')

for col in dtype_cols:
    if not pd.api.types.is_datetime64_dtype(hist_tags_org_df[col]):
        logging.error(f"kolom '{col}' in '{hist_tags_csv}' kan niet worden geconverteerd"
                      " naar np.datetime64 formaat. Controleer of deze datums realistisch zijn.")
        sys.exit()

#%% filteren hist_tags op alles wat niet in ignored staat
if mpt_ignore_csv:
    logging.info(f'histag_ignore wordt gelezen uit {mpt_ignore_csv.absolute().resolve()}')
    consistency_df['histTag_ignore'] = pd.read_csv(mpt_ignore_csv,sep=None,header=0,engine='python') 
else:
    logging.info(f'histag_ignore wordt gelezen uit werkblad "histTag_ignore" in {consistency_in.absolute().resolve()}')
consistency_df['histTag_ignore']['UNKNOWN_SERIE'] = consistency_df['histTag_ignore']['UNKNOWN_SERIE'].str.replace('#','')  

hist_tags_df = hist_tags_org_df.copy()
hist_tags_df['fews_locid'] = hist_tags_org_df.apply(idmap2tags, args=[idmap_total], axis=1)
hist_tags_no_match_df = hist_tags_df[hist_tags_df['fews_locid'].isna()]
hist_tags_no_match_df = hist_tags_no_match_df[~hist_tags_no_match_df['serie'].isin(consistency_df['histTag_ignore']['UNKNOWN_SERIE'])] 
hist_tags_no_match_df = hist_tags_no_match_df.drop('fews_locid',axis=1)
hist_tags_no_match_df.columns = ['UNKNOWN_SERIE','STARTDATE','ENDDATE']
hist_tags_no_match_df = hist_tags_no_match_df.set_index('UNKNOWN_SERIE')
consistency_df['histTags_noMatch'] = hist_tags_no_match_df

if not consistency_df['histTags_noMatch'].empty:
    logging.warning('{} histTags zijn niet opgenomen in idmap'.format(len(consistency_df['histTags_noMatch'])))
else:
    logging.info('alle histTags zijn opgenomen in idmap')

#%% wegschrijven van ids die ten onrechte in ignore-lijst staan
hist_tags_opvlwater_df = hist_tags_org_df.copy()
hist_tags_opvlwater_df['fews_locid'] = hist_tags_org_df.apply(idmap2tags, args=[idmap_dict['IdOPVLWATER']], axis=1)
hist_tags_opvlwater_df = hist_tags_opvlwater_df[hist_tags_opvlwater_df['fews_locid'].notna()]
hist_tag_ignore_match_df = consistency_df['histTag_ignore'][consistency_df['histTag_ignore']['UNKNOWN_SERIE'].isin(hist_tags_opvlwater_df['serie'])]
hist_tag_ignore_match_df = hist_tag_ignore_match_df.set_index('UNKNOWN_SERIE')
consistency_df['histTags_ignore_match'] = hist_tag_ignore_match_df

if not consistency_df['histTags_ignore_match'].empty:
    logging.warning('{} histTags zijn ten onrechte opgenomen in histTag ignore'.format(len(consistency_df['histTags_ignore_match'])))
else:
    logging.info('geen histTags ten onrechte in ignore')

#%% aanmaken van mpt_df vanuit de fews_locid lijsten in hist_tags_df
logging.info('omzetten van histTags naar meetpunten')
hist_tags_df = hist_tags_df[hist_tags_df['fews_locid'].notna()]
mpt_hist_tags_df = hist_tags_df.explode('fews_locid').reset_index(drop=True)

# bepalen minimale start en maximale eindtijd per fews_locid. 
mpt_df = pd.concat([mpt_hist_tags_df.groupby(['fews_locid'], sort=False)['total_min_start_dt'].min(),
                    mpt_hist_tags_df.groupby(['fews_locid'], sort=False)['total_max_end_dt'].max()],
                   axis=1)

mpt_df = mpt_df.sort_index(axis=0)
mpt_df.columns = ['STARTDATE','ENDDATE']
mpt_df.index.name = 'LOC_ID'

# alle hoofdloc waar geen histag op binnekomt toevoegen
kw_locs = list(mpt_df[mpt_df.index.str.contains('KW', regex=False)].index)
h_locs = np.unique(['{}0'.format(loc[0:-1]) for loc in kw_locs])
h_locs_missing = [loc for loc in h_locs if not loc in list(mpt_df.index)]
h_locs_df = pd.DataFrame(data={'LOC_ID' : h_locs_missing,
                               'STARTDATE' : [pd.NaT]*len(h_locs_missing),
                               'ENDDATE' :  [pd.NaT]*len(h_locs_missing)})
h_locs_df = h_locs_df.set_index('LOC_ID')

mpt_df = pd.concat([mpt_df,h_locs_df],axis=0)
# de start en eindtijd op de hoofdlocatie updaten met de min/max van de sublocatie

mpt_df[['STARTDATE','ENDDATE']] = mpt_df.apply(update_hlocs,axis=1,result_type="expand")

mpt_df = mpt_df.sort_index()
consistency_df['mpt'] = mpt_df

#%% consistentie parameters: zijn alle interne parameters opgenomen in parameters.xml
logging.info('controle dubbele idmaps')
consistency_df['dubbele idmaps'] = pd.DataFrame(columns=['bestand',
                                                    'externalLocation',
                                                    'externalParameter',
                                                    'internalLocation',
                                                    'internalParameter'])                    
for idmap_file in idmap_files:
    idmap_doubles = [id_map for id_map in idmap_dict[idmap_file] if idmap_dict[idmap_file].count(id_map) > 1]
    if len(idmap_doubles) > 0:
        idmap_doubles = list({idmap['externalLocation']:idmap for idmap in idmap_doubles}.values())
        df = pd.DataFrame(idmap_doubles,columns=['internalLocation','externalLocation','internalParameter','externalParameter'])
        df['bestand'] = idmap_file
        consistency_df['dubbele idmaps'] = pd.concat([consistency_df['dubbele idmaps'], df], axis=0)
        logging.warning('{} dubbele idmap(s) in {}'.format(len(idmap_doubles),idmap_file))
    else:
        logging.info('geen dubbele idmaps in {}'.format(idmap_file))

#%% consistentie parameters: zijn alle interne parameters opgenomen in parameters.xml
logging.info('zoeken op missende interne parameters')
config_parameters = list(fews_config.get_parameters(dict_keys='parameters').keys())
id_map_parameters = [id_map['internalParameter'] for id_map in idmap_total]
params_missing = [parameter for parameter in id_map_parameters 
                  if not parameter in config_parameters]

if len(params_missing) == 0:
    logging.info('alle parameters in idMaps zijn opgenomen in config')
else:
    logging.warning('{} parameter(s) in idMaps missen in config'.format(len(params_missing)))
    consistency_df['params_missing'] =  pd.DataFrame({'parameters': params_missing})
    consistency_df['params_missing'] = consistency_df['params_missing'].set_index('parameters')

#%% controle op consistentie sublocs t.b.v. wegschrijven hoofdloc_gdf
logging.info('controle consistentie sublocs op per hoofdlocatie')

if 'xy_ignore' in consistency_df.keys():
    xy_ignore_df = consistency_df['xy_ignore']
else:
    xy_ignore_df = pd.DataFrame({'internalLocation':[],'x':[],'y':[]})
    
hoofdloc_gdf = fews_config.get_locations('OPVLWATER_HOOFDLOC')
subloc_gdf = fews_config.get_locations('OPVLWATER_SUBLOC')

hloc_errors = {'LOC_ID':[],
               'SUB_LOCS':[],
               'LOC_NAME':[],
               'GEOMETRY':[],
               'SYSTEEM':[],
               'RAYON':[],
               'KOMPAS':[]}

grouper = subloc_gdf.groupby('PAR_ID')
par_dict = {'LOC_ID':[],
            'LOC_NAME':[],
            'X':[],
            'Y':[],
            'ALLE_TYPES':[],
            'START':[],
            'EIND':[],
            'SYSTEEM':[],
            'RAYON':[],
            'KOMPAS':[]}

for loc_id, gdf in grouper:
    caw_code = loc_id[2:-2]
    errors = dict.fromkeys(['LOC_NAME','GEOMETRY','SYSTEEM','RAYON','KOMPAS'],False)
    fields = dict.fromkeys(par_dict.keys(),None)
    fields['LOC_ID'] = loc_id
    # controle subloc op 1 consistente parent sub-string
    loc_names = np.unique(gdf['LOC_NAME'].str.extract(pat = f'([A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*)').values)
    if not len(loc_names) == 1:
        errors['LOC_NAME'] = ",".join(loc_names)
    else:
        fields['LOC_NAME'] = loc_names[0]
    #controle subloc op 1 consistente locatie
    if any([re.match(loc,loc_id) for loc in xy_ignore_df['internalLocation']]):
        fields['X'], fields['Y'] = next([row['x'],row['y']] 
                                        for index, row in xy_ignore_df.iterrows() 
                                        if re.match(row['internalLocation'], loc_id))
    else:
        geoms = gdf['geometry'].unique()
        if not len(geoms) == 1:
            errors['GEOMETRY'] = ",".join([f'({geom.x} {geom.y})' for geom in geoms])
        else:
            fields['X'] = geoms[0].x
            fields['Y'] = geoms[0].y
        
    #wegschrijven alle types op sublocaties
    all_types = list(gdf['TYPE'].unique())
    all_types.sort()
    fields['ALLE_TYPES'] = '/'.join(all_types)
    
    #wegschrijven start/eind uit min/max sublocaties
    fields['START'] = gdf['START'].min()
    fields['EIND'] = gdf['EIND'].max()
    
    #controle op unieke atributen
    for attribuut in ['SYSTEEM','RAYON','KOMPAS']:
        vals = gdf[attribuut].unique()
        if not len(vals) == 1:
            errors[attribuut] = "","".join(vals)
        else:
            fields[attribuut] = vals[0]     

    # parent kan geschreven worden als alle subloc-waarden consistent zijn
    if not None in fields.values():
        for key,value in fields.items():
            par_dict[key].append(value)
            
    # als fout, dan opname in error-dict
    if any(errors.values()):
        hloc_errors['LOC_ID'].append(loc_id)
        hloc_errors['SUB_LOCS'].append(','.join(gdf['LOC_ID'].values))
        for key,value in errors.items():
            if value == False:
                value = ''
            hloc_errors[key].append(value)

consistency_df['hloc error'] = pd.DataFrame(hloc_errors)
#opname in samenvatting

if consistency_df['hloc error'].empty:
    logging.info('geen fouten in aanmaken hoofdlocaties')
    par_gdf = pd.DataFrame(par_dict)
    columns = list(hoofdloc_gdf.columns)
    drop_cols = [col for col in hoofdloc_gdf.columns if (col in par_gdf.columns) & (not col =='LOC_ID')]
    drop_cols = drop_cols + ['geometry']
    hoofdloc_gdf = hoofdloc_gdf.drop(drop_cols, axis=1)
    hoofdloc_gdf = par_gdf.merge(hoofdloc_gdf,on='LOC_ID')
    hoofdloc_gdf['geometry'] = hoofdloc_gdf.apply((lambda x: Point(float(x['X']),
                                                                   float(x['Y']))),
                                                                    axis=1)
    hoofdloc_gdf = hoofdloc_gdf[columns]
    
    
else:
    logging.warning('{} fouten bij aanmaken hoofdlocaties'.format(len(consistency_df['hloc error'])))
    logging.warning('hoofdlocaties worden alleen opnieuw geschreven vanuit sublocaties wanneer de fouten zijn opgelost')  



#%% consistentie externe parameters met interne parameters/locaties
logging.info('controle foutieve ex-parameters & niet opgenomen inlocs')

if 'externalParametersAllowed' in config.keys():
    expars_allowed = {key: value.replace(" ","").split(',') 
                      for key, value in config['externalParametersAllowed'].items()}
    
waterstandloc_gdf = fews_config.get_locations('OPVLWATER_WATERSTANDEN_AUTO')
mswloc_gdf = fews_config.get_locations('MSW_STATIONS')

ex_par_errors = {'internalLocation':[],
                 'locationType':[],
                 'exParError':[],
                 'types':[],
                 'FQ':[],
                 'I.X':[],
                 'IX.':[],
                 'SS./SM.':[]}

int_loc_missing = []

#maak een data-frame zodat we kunnen groeperen bij internalLocation
idmap_df = pd.DataFrame.from_dict(idmap_dict['IdOPVLWATER'])

for int_loc, loc_group in idmap_df.groupby('internalLocation'):
    #initieer een aantal variabelen
    
    errors = dict.fromkeys(['I.X','IX.','FQ', 'SS./SM.'],False)
    
    #interne locatie en externe parameters
    ex_pars = np.unique(loc_group['externalParameter'].values)
    ex_pars_gen = [re.sub("\d", ".", ex_par) for ex_par in ex_pars]
    
    #vaststellen locatie-type
    if int_loc in hoofdloc_gdf['LOC_ID'].values:
        loc_properties = hoofdloc_gdf[hoofdloc_gdf['LOC_ID'] == int_loc]
        loc_type = 'hoofdloc'
    elif int_loc in subloc_gdf['LOC_ID'].values:
        loc_properties = subloc_gdf[subloc_gdf['LOC_ID'] == int_loc]
        loc_type = 'subloc'
        regexes = ['HR.$']
    elif int_loc in waterstandloc_gdf['LOC_ID'].values:
        loc_type = 'waterstandloc'
    elif int_loc in mswloc_gdf['LOC_ID'].values:
        loc_type = 'mswloc'
    else:
        loc_type = None
        int_loc_missing += [int_loc]
    
    #vaststellen object_typen
    if loc_type in ['hoofdloc', 'subloc']:
        all_types = loc_properties['ALLE_TYPES'].values[0].split("/")
        all_types = [item.lower() for item in all_types]
    elif loc_type == 'waterstandloc':
        all_types = ['waterstandloc']
    
    if loc_type == 'subloc':
        sub_type = subloc_gdf[subloc_gdf['LOC_ID'] == int_loc]['TYPE'].values[0]
        
        #zoeken naar foutief toegekende ex_pars
        regexes += [j for i in
                    [values for keys, values in expars_allowed.items() if keys in all_types]
                    for j in i]
        regexes += list(dict.fromkeys(regexes))
        
        ex_par_error = [ex_par for ex_par in ex_pars if not any([regex.match(ex_par) for regex in [re.compile(rex) for rex in regexes]])]
        
        # als sub_type = schuif dan SM. of SS.
        if sub_type == 'schuif':
            if not any([ex_par for ex_par in ex_pars_gen if ex_par in ['SS.', 'SM.']]):
                errors['SS./SM.'] = True
        
        # als wel/niet I.B dan ook wel/niet IB.
        if any([ex_par for ex_par in ex_pars_gen if ex_par in ['I.B', 'I.H', 'I.L']]):
            if not any([ex_par for ex_par in ex_pars_gen if ex_par in ['IB.', 'IH.', 'IL.']]):
                errors['IX.'] = True
        elif any([ex_par for ex_par in ex_pars_gen if ex_par in ['IB.', 'IH.', 'IL.']]):
             errors['I.X'] = True
        
        # Als FQ, dan ook I.B.
        if 'FQ.' in ex_pars_gen: 
            if not any([ex_par for ex_par in ex_pars_gen if ex_par in ['IB.', 'IH.', 'IL.', 'I.B', 'I.H', 'I.L']]):
                errors['FQ'] = True
                
    elif loc_type == 'hoofdloc':
        
        #zoeken naar foutief toegekende ex_pars
        regexes = ['HS.$', 'QR.$', 'QS.$', 'WR', 'WS']
        
        ex_par_error = [ex_par for ex_par in ex_pars if not any([regex.match(ex_par) for regex in [re.compile(rex) for rex in regexes]])]
                   
    else:
        ex_par_error = []
            
    # rapporteren expar_errors
    if len(ex_par_error) > 0 | any(errors.values()):
        ex_par_errors['internalLocation'].append(int_loc)
        ex_par_errors['locationType'].append(loc_type)
        ex_par_errors['exParError'].append(','.join(ex_par_error))
        ex_par_errors['types'].append(','.join(all_types))
        for key, value in errors.items():
            ex_par_errors[key].append(value)
        
    
#opname in data-frame           
consistency_df['exPar error'] = pd.DataFrame(ex_par_errors)
consistency_df['intLoc missing'] = pd.DataFrame({'internalLocation':int_loc_missing})

#opname in samenvatting

#loggen van resultaat
if len(consistency_df['exPar error']) == 0:
    logging.info('geen ExPar errors')
else:
  logging.warning('{} locaties met ExPar errors'.format(len(consistency_df['exPar error'])))
      
if len(consistency_df['intLoc missing']) == 0:
    logging.info('alle interne locaties uit idmap opgenomen in locationSets')
else:
    logging.warning('{} interne locaties niet opgenomen in locationSets'.format(len(consistency_df['intLoc missing'])))                    

#%% expar missings
logging.info('controle missende ex-parameters')
ex_par_missing = {'internalLocation':[],
                  'exPars':[],
                  'QR':[],
                  'QS':[],
                  'HS':[]}

grouper = idmap_df.groupby('internalLocation')
for index, row in hoofdloc_gdf.iterrows():
    missings = dict.fromkeys(['QR','QS','HS'],False)
    int_loc = row['LOC_ID']
    
    loc_group = next((df for loc,df in idmap_df.groupby('internalLocation') if loc == int_loc), pd.DataFrame())
    
    if not loc_group.empty:
        ex_pars = np.unique(loc_group['externalParameter'].values)
        ex_pars_gen = [re.sub("\d", ".", ex_par) for ex_par in ex_pars]
    else:
        ex_pars = []
        ex_pars_gen = []
    
    #is er een HS?
    if not ('HS.' in ex_pars_gen):
        missings['HS'] = True
        
    if not ('QR.' in ex_pars_gen):
        missings['QR'] = True
        
    if not ('QS.' in ex_pars_gen):
        missings['QS'] = True
        
    # rapporteren missings
    if any(missings.values()):
        ex_par_missing['internalLocation'].append(int_loc)
        ex_par_missing['exPars'].append(','.join(ex_pars))
        for key, value in missings.items():
            ex_par_missing[key].append(value)

consistency_df['exPar missing'] = pd.DataFrame(ex_par_missing)

#loggen van resultaat
if len(consistency_df['exPar missing']) == 0:
    logging.info('geen ExPar missing')
else:
  logging.warning('{} locaties met ExPar missing'.format(len(consistency_df['exPar missing'])))

#%% zoeken naar ex-loc errors
logging.info('controle externe locaties')
ex_loc_errors = {'internalLocation':[],
                 'externalLocation':[]}   

for loc_group in idmap_df.groupby('externalLocation'):

    #initialiseren int_loc_error
    int_loc_error = []
    
    #zoeken naar ex-loc errors
    ex_loc = loc_group[0]
    int_locs = np.unique(loc_group[1]['internalLocation'].values)
    
    # als lengte van ex-loc == 3
    if len(ex_loc) == 3:
        
        # de default-case
        if not bool(re.match('8..$',ex_loc)):
            int_loc_error = [int_loc for int_loc in int_locs if 
                             not bool(re.match(f'...{ex_loc}..$',int_loc))]
         
        # opgesplitste locaties; ex-loc altijd naar 1 unieke hoofdlocatie + subloc
        else:
            for loc_type in ['KW','OW']:
                int_locs_select = [int_loc for int_loc in int_locs 
                                   if bool(re.match(f'{loc_type}.',int_loc))]
                if len(np.unique([int_loc[:-1] for int_loc in int_locs_select])) > 1:
                    int_loc_error += list(int_locs_select)
    
    # als lengte ex-loc == 4
    if len(ex_loc) == 4:
        
        # de default-case
        if not bool(re.match('.8..$',ex_loc)):
            int_loc_error += [int_loc for int_loc in int_locs if 
                              not bool(re.match(f'..{ex_loc}..$',int_loc))]
        
        # opgesplitste locaties; ex-loc altijd naar 1 unieke hoofdlocatie + subloc
        else:
            for loc_type in ['KW','OW']:
                int_locs_select = [int_loc for int_loc in int_locs 
                                   if bool(re.match(f'{loc_type}.',int_loc))]
                if len(np.unique([int_loc[:-1] for int_loc in int_locs_select])) > 1:
                    int_loc_error += list(int_locs_select)
    
    #als de ex-loc in de ignore-lijst staan, dan int_loc_error opruimen
    if 'exLoc_ignore' in consistency_df.keys():
        if int(ex_loc) in consistency_df['exLoc_ignore']['externalLocation'].values:
            int_loc_error = [int_loc for int_loc in int_loc_error 
                               if not int_loc in 
                               consistency_df['exLoc_ignore'][consistency_df['exLoc_ignore']['externalLocation'] 
                                                         == int(ex_loc)]['internalLocation'].values]
             
    for int_loc in int_loc_error:
        ex_loc_errors['internalLocation'].append(int_loc)
        ex_loc_errors['externalLocation'].append(ex_loc)

consistency_df['exLoc error'] = pd.DataFrame(ex_loc_errors)

if len(consistency_df['exLoc error']) == 0:
    logging.info('alle externe locaties consistent met interne locaties')
else:
    logging.warning('{} externe locaties onlogisch bij interne locaties'.format(len(consistency_df['exLoc error'])))                    

#%% zoeken naar sub-locaties anders dan krooshek en debietmeter:
#   - zonder stuurpeil tijdserie
#   - waarbij meerdere tijdseries met stuurpeilen naar dezelfde interne paramer mappen
logging.info('controle koppeling tijdseries')
if 'TS800_ignore' in consistency_df.keys():
    ts_ignore_df = consistency_df['TS800_ignore']
else:
    ts_ignore_df = pd.DataFrame({'internalLocation':[],'externalLocation':[]})

idmap_subloc_df = idmap_df[idmap_df['internalLocation'].isin(subloc_gdf['LOC_ID'].values)] # alleen locaties die in de sub-locs locationSet zitten
idmap_subloc_df['type'] = idmap_subloc_df['internalLocation'].apply((lambda x: subloc_gdf[subloc_gdf['LOC_ID'] == x]['TYPE'].values[0])) #toevoegen van type
idmap_subloc_df['loc_groep'] = idmap_subloc_df['internalLocation'].apply((lambda x: x[0:-1]))

ts_errors = {'internalLocation':[],
             'internalParameters':[],
             'externalParameters':[],
             'externalLocations':[],
             'type':[],
             'fout':[]
             }

for loc_group, group_df in idmap_subloc_df.groupby('loc_groep'):
    
    #uniek nummer per ex-loc
    ex_locs = np.unique(group_df['externalLocation'].values)
    ex_locs_dict = {ex_loc:idx for idx, ex_loc in enumerate(ex_locs)}
    
    #vinden van 800 nummers
    split_ts = [key for key in ex_locs_dict.keys() if 
                  any([regex.match(key) 
                        for regex in [re.compile(rex) 
                                      for rex in ['8..','.8..']]])]
    
    ex_locs_skip = ts_ignore_df[ts_ignore_df['internalLocation'].isin(group_df['internalLocation'])]['externalLocation']
    
    split_ts = [key for key in split_ts if not str(key) in ex_locs_skip.values.astype(np.str)]
    
    ex_locs_dict = {k:(ex_locs_dict[k[1:]] 
                        if (k[1:] in ex_locs_dict.keys()) and (not k in split_ts) 
                        else v) for (k,v) in ex_locs_dict.items()}
    
    org_uniques = np.unique([val for key,val in ex_locs_dict.items() if not key in split_ts])
       
    # als er maar 1 groep zit in split_ts én een groep in de originele tijdseriegroepen, dan samenvoegen
    if (len(org_uniques) == 1) & (len(split_ts) == 1):
        ex_locs_dict = {k:(org_uniques[0] if k in split_ts else v) for (k,v) in ex_locs_dict.items()}
        
    group_df['ex_loc_group'] = group_df['externalLocation'].apply((lambda x: ex_locs_dict[x]))
 
    for int_loc, loc_df in group_df.groupby('internalLocation'):
        sub_type = subloc_gdf[subloc_gdf['LOC_ID'] == int_loc]['TYPE'].values[0]
        end_time = pd.to_datetime(subloc_gdf[subloc_gdf['LOC_ID'] == int_loc]['EIND'].values[0])
        ex_pars = np.unique(loc_df['externalParameter'].values)
        int_pars = np.unique(loc_df['internalParameter'].values)
        ex_locs = np.unique(loc_df['externalLocation'].values)
        
        
        if sub_type in ['krooshek','debietmeter']:
            if any([re.match('HR.',ex_par) for ex_par in ex_pars]):
                #krooshek/debietmeter met stuurpeil = fout
                ts_errors['internalLocation'].append(int_loc)
                ts_errors['internalParameters'].append(",".join(int_pars))
                ts_errors['externalParameters'].append(",".join(ex_pars))
                ts_errors['externalLocations'].append(','.join(ex_locs))
                ts_errors['type'].append(sub_type)
                ts_errors['fout'].append(f'{sub_type} met stuurpeil')
        
        else: #geen krooshek of debietmeter
            # geen sp, maar wel sp op andere subloc = fout
            if (not any([re.match('HR.',ex_par) for ex_par in ex_pars])): # geen stuurpeil
                if any([re.match('HR.',ex_par) for ex_par in np.unique(group_df['externalParameter'])]):
                    #~krooshek/debietmeter zonder stuurpeil = fout
                    if not sub_type in ['totaal', 'vispassage']:
                        if pd.Timestamp.now() < end_time:
                            sp_locs = np.unique(group_df[group_df['externalParameter'].str.match('HR.')]['internalLocation'])
                            ts_errors['internalLocation'].append(int_loc)
                            ts_errors['internalParameters'].append(",".join(int_pars))
                            ts_errors['externalParameters'].append(",".join(ex_pars))
                            ts_errors['externalLocations'].append(','.join(ex_locs))
                            ts_errors['type'].append(sub_type)
                            ts_errors['fout'].append(f'{sub_type} zonder stuurpeil ({",".join(sp_locs)} wel)')
                    
            else: #krooshek/debietmeter met stuurpeil
                # >1 sp zonder andere interne parameter = fout
                time_series = loc_df.groupby(['ex_loc_group','externalParameter'])
                sp_series = [series for series in time_series if bool(re.match('HR.',series[0][1]))]
                for idx, series in enumerate(sp_series):
                    ex_par = series[0][1]
                    ex_locs = series[1]['externalLocation']
                    int_par = np.unique(series[1]['internalParameter'])
                    if len(int_par) > 1:
                        # 1 sp series gekoppeld aan 2 fews parameters
                        ts_errors['internalLocation'].append(int_loc)
                        ts_errors['internalParameters'].append(",".join(int_pars))
                        ts_errors['externalParameters'].append(",".join(ex_pars))
                        ts_errors['externalLocations'].append(','.join(ex_locs))
                        ts_errors['type'].append(sub_type)
                        ts_errors['fout'].append(f'{",".join(int_par)} gekoppeld aan 1 sp-serie (exPar: {ex_par}, exLoc(s)): {",".join(ex_locs)}')
                        
                    other_series = [series for idy, series in enumerate(sp_series) if not idy == idx]
                    other_int_pars = [np.unique(series[1]['internalParameter']) for series in other_series]
                    if len(other_int_pars) > 0: other_int_pars = np.concatenate(other_int_pars)
                    conflicting_pars = [par for par in int_par if par in other_int_pars]
                    if len(conflicting_pars) > 0:
                        # 2 sp series gekoppeld aan dezelfde fews parameter
                        ts_errors['internalLocation'].append(int_loc)
                        ts_errors['internalParameters'].append(",".join(int_pars))
                        ts_errors['externalParameters'].append(",".join(ex_pars))
                        ts_errors['externalLocations'].append(','.join(ex_locs))
                        ts_errors['type'].append(sub_type)
                        ts_errors['fout'].append(f'{",".join(conflicting_pars)} gekoppeld aan sp-serie (exPar: {ex_par}, exLoc(s)): {",".join(ex_locs)}')
                               
consistency_df['timeSeries error'] = pd.DataFrame(ts_errors)

#opname in samenvatting
if len(consistency_df['timeSeries error']) == 0:
    logging.info('alle tijdseries zijn logisch gekoppeld aan interne locaties/parameters')
else:
    logging.warning('{} tijdseries missend/onlogisch gekoppeld'.format(len(consistency_df['timeSeries error'])))   

#%% controle validationrulesets
'''ToDo:
    - df_idmap moet IDoppervlaktewater + Hymos bevatten
    - kolom fout_type en fout_opmerking toevoegen/splitsen
    - internal parameters uniek maken
    - fout melden als smin/smax lijsten niet unieke waarden bevatten
    - comment regel in ini-file bij lijst-vergelijkingen. Herstructureren json zodat die fout niet gemaakt kan worden?
'''
logging.info('controle validationRules')

valid_errors = {'internalLocation':[],
                'start':[],
                'eind':[],
                'internalParameters':[],
                'fout_type':[],
                'fout_beschrijving':[]
                }

def sort_attribs(rule):
    result = {}
    for key,value in rule.items():
        if isinstance(value,str):
            result[key] = [value]
        elif isinstance(value,list):
            periods = [val['period'] for val in value]
            attribs = [val['attribute'] for val in value]
            result[key] = [attrib for _,attrib in sorted(zip(periods,attribs))]
    return result
            
            

location_sets_dict = xml_to_dict(fews_config.RegionConfigFiles['LocationSets'])['locationSets']['locationSet']
for set_name in config['validation_rules'].keys():
    #set_name = 'subloc'
    location_set = location_sets[set_name]
    location_set_meta = next(loc_set for loc_set in location_sets_dict if loc_set['id'] == location_set['id'])['csvFile']
    location_set_gdf = location_set['gdf']
    
    attrib_files = location_set_meta['attributeFile']
    
    if not isinstance(attrib_files,list):
        attrib_files = [attrib_files]
    
    attrib_files = [attrib_file for attrib_file in attrib_files if 'attribute' in attrib_file.keys()]
    
    for attrib_file in attrib_files:    
        # schone lijst met attributen verkrijgen
        attribs = attrib_file['attribute']
        join_id =  attrib_file['id'].replace("%","")
        if not isinstance(attrib_file['attribute'],list):
            attribs = [attribs]        
        attribs = [attrib['number'].replace("%",'') for attrib in attribs if 'number' in attrib.keys()]
        
        # attribuut-bestand relateren op locatie aan locationSet
        attrib_df = pd.read_csv(fews_config.MapLayerFiles[attrib_file['csvFile'].replace('.csv','')],
                                sep=None,
                                engine='python')
        
        attrib_df.rename(columns={join_id:'LOC_ID'},inplace=True)
        drop_cols = [col for col in attrib_df if not col in attribs + ['LOC_ID']]
        attrib_df = attrib_df.drop(columns=drop_cols, axis=1)
    
        location_set_gdf = location_set_gdf.merge(attrib_df,
                                                  on='LOC_ID',
                                                  how='outer')
        
    validation_rules = config['validation_rules'][set_name]
    validaton_attributes = get_attribs(validation_rules)
    
    #row = location_set_gdf.loc[0]
    params_df = pd.DataFrame.from_dict({int_loc:[df['internalParameter'].values] 
                                        for int_loc, df 
                                        in idmap_df.groupby('internalLocation')}, 
                                       orient='index', 
                                       columns=['internalParameters'])
    
    for (idx, row) in location_set_gdf.iterrows():
        int_loc = row['LOC_ID']
        row = row.dropna()
        if set_name == 'sublocaties':
            loc_type = row['TYPE']
        
        if int_loc in params_df.index:
            int_pars = np.unique(params_df.loc[int_loc]['internalParameters'])
        else:
            int_pars = []
            
        attribs_required = get_attribs(validation_rules,int_pars)
        attribs_missing = [attrib for attrib in attribs_required if not attrib in row.keys()]
        attribs_obsolete = [attrib for attrib in validaton_attributes if 
                            (not attrib in attribs_required) and (attrib in row.keys())]
        attribs = [attrib for attrib in attribs_required if not attrib in attribs_missing]
        
        for key, value in {'missend':attribs_missing,'overbodig':attribs_obsolete}.items():
            if len(value) > 0:
                valid_errors['internalLocation'] += [int_loc]
                valid_errors['start'] += [row['START']]
                valid_errors['eind'] += [row['EIND']]
                valid_errors['internalParameters'] += [",".join(int_pars)]
                valid_errors['fout_type'] += [key]
                valid_errors['fout_beschrijving'] += [",".join(value)]
            
        for validation_rule in validation_rules:
            errors = {'fout_type':None,
                      'fout_beschrijving':[]}
            
            param = validation_rule['parameter']
            if any(re.match(param,int_par) for int_par in int_pars):
                rule = validation_rule['extreme_values']
                rule = sort_attribs(rule)
                #regels met alleen hmax/hmin
                if all(key in ['hmax', 'hmin'] for key in rule.keys()):
                    for hmin, hmax in zip(rule['hmin'], rule['hmax']):
                        if all(attrib in row.keys() for attrib in [hmin, hmax]):
                            if row[hmax] < row[hmin]:
                                errors['fout_type'] = 'waarde'
                                errors['fout_beschrijving'] += [f"{hmax} < {hmin}"]
                #regels met soft + hard min/max
                elif all(key in rule.keys() for key in ['hmax', 'smax', 'smin', 'hmin']):
                    hmax = rule['hmax'][0]
                    hmin = rule['hmin'][0]
                    for smin, smax in zip(rule['smin'], rule['smax']):
                        if all(attrib in row.keys() for attrib in [smin, smax]):
                            if row[smax] <= row[smin]:
                                errors['fout_type'] = 'waarde'
                                errors['fout_beschrijving'] += [f"{smax} <= {smin}"]
                            if row[hmax] < row[smax]:
                                errors['fout_type'] = 'waarde'
                                errors['fout_beschrijving'] += [f"{'hmax'} < {smax}"]
                            if row[smin] < row[hmin]:
                                errors['fout_type'] = 'waarde'
                                errors['fout_beschrijving'] += [f"{smin} < {hmin}"]
                                
            valid_errors['internalLocation'] += [row['LOC_ID']] * len(errors['fout_beschrijving'])
            valid_errors['start'] += [row['START']] * len(errors['fout_beschrijving'])
            valid_errors['eind'] += [row['EIND']] * len(errors['fout_beschrijving'])
            valid_errors['internalParameters'] += [",".join(int_pars)] * len(errors['fout_beschrijving'])
            valid_errors['fout_type'] += [errors['fout_type']] * len(errors['fout_beschrijving'])
            valid_errors['fout_beschrijving'] += errors['fout_beschrijving']
    
consistency_df['validation error'] = pd.DataFrame(valid_errors)
consistency_df['validation error'] = consistency_df['validation error'].drop_duplicates()

#opname in samenvatting
if len(consistency_df['validation error']) == 0:
    logging.info('er zijn geen foute/missende validatieregels')
else:
    logging.warning('{} validatieregels zijn fout/missend'.format(len(consistency_df['validation error'])))

#%% controle op expar
logging.info('regex externalParameters')
par_errors = {'internalLocation':[],
             'internalParameter':[],
             'externalParameter':[],
             'fout':[]
             }

internal_parameters = [mapping['internal'] for mapping in config['parameter_mapping']]
for idx, row in idmap_df.iterrows():
    error = None
    ext_par = None
    ext_par = [mapping['external'] for mapping in config['parameter_mapping'] if 
                    re.match(f'{mapping["internal"]}[0-9]',row['internalParameter'])]
    
    if ext_par:
        if not any(re.match(par,row['externalParameter']) for par in ext_par):
            error = 'parameter mismatch'
    
    else:
        error = 'pars niet opgenomen in config'
        
    if error:
        par_errors['internalLocation'].append(row['internalLocation'])
        par_errors['internalParameter'].append(row['internalParameter'])
        par_errors['externalParameter'].append(row['externalParameter'])
        par_errors['fout'].append(error)

consistency_df['par mismatch'] = pd.DataFrame(par_errors)
#opname in samenvatting

if len(consistency_df['par mismatch']) == 0:
    logging.info('geen regex fouten op interne en externe parameters')
else:
    logging.warning('{} regex fouten op interne en externe parameters'.format(len(consistency_df['par mismatch'])))    

#%% validatie locationSets
logging.info('validatie locatieSets')
loc_set_errors = {'locationId':[],
                  'caw_code':[],
                  'caw_name':[],
                  'csv_file':[],
                  'location_name':[],
                  'type':[],
                  'functie': [],
                  'name_error':[],
                  'caw_name_inconsistent':[],
                  'missing_in_map':[],
                  'missing_in_set':[],
                  'missing_peilschaal':[],
                  'missing_hbov':[],
                  'missing_hben':[],
                  'missing_hbovps':[],
                  'missing_hbenps':[],
                  'missing_hloc':[],
                  'xy_not_same':[]}

sets = {'waterstandlocaties':'WATERSTANDLOCATIES',
        'sublocaties': 'KUNSTWERKEN',
        'hoofdlocaties': 'KUNSTWERKEN'}

'''
ToDo:
    missing_in_set: 
        - kijken in welke set een locatie hoort te zitten en of dat klopt met de sectie?
        - kijken of de locatie mist in de locatieset. Even een locatie uit de csv halen om te testen.
'''
    
for set_name,section_name in sets.items():
    logging.info(set_name)
    location_set = location_sets[set_name]
    location_gdf = location_set['gdf']
    csv_file = fews_config.locationSets[location_set['id']]['csvFile']['file']
    int_locs = [] 

    for idmap in ['IdOPVLWATER', 'IdOPVLWATER_HYMOS']:
        for section in idmap_sections[idmap][section_name]: 
            int_locs += [item['internalLocation'] for item in xml_to_dict(fews_config.IdMapFiles[idmap],**section)['idMap']['map']]
    
    if set_name == 'sublocaties':
        int_locs = [loc for loc in int_locs if not loc[-1] == '0']
        par_gdf = location_sets['hoofdlocaties']['gdf']
        
    elif set_name == 'hoofdlocaties':
        int_locs = [loc for loc in int_locs if loc[-1] == '0']
    
    #idx, row = list(location_gdf.iterrows())[0]
    for idx, row in list(location_gdf.iterrows()):
        error = {'name_error':False,
                 'caw_name_inconsistent':False,
                 'missing_in_map':False,
                 'type':'',
                 'functie':'',
                 'missing_in_set':False,
                 'missing_peilschaal':False,
                 'missing_hbov':False,
                 'missing_hben':False,
                 'missing_hbovps':False,
                 'missing_hbenps':False,
                 'missing_hloc':False,
                 'xy_not_same':False}
        
        loc_id = row['LOC_ID']
        caw_code = loc_id[2:-2]
        loc_name = row['LOC_NAME']
        caw_name = ''  
        
        if set_name == 'sublocaties':
            
            loc_functie = row['FUNCTIE']
            sub_type = row['TYPE']
            
            if sub_type in ['afsluiter', 'debietmeter', 'krooshek', 'vispassage']:
                if not re.match(f'[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*-{sub_type}',loc_name):
                    error['name_error'] = True

            else:
                if not re.match(f'[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*-{sub_type}[0-9]_{loc_functie}',loc_name):
                    error['name_error'] = True
            
            if not error['name_error']:
                caw_name = re.match(f'([A-Z0-9 ]*)_',loc_name).group(1)
                if not all(location_gdf[location_gdf['LOC_ID'].str.match(
                        f'..{caw_code}')]['LOC_NAME'].str.match(
                            f'({caw_name}_{caw_code}-K)')
                            ):
                                error['caw_name_inconsistent'] = True
                
            if not row['HBOV'] in location_sets['waterstandlocaties']['gdf']['LOC_ID'].values:
                error['missing_hbov'] = True
            
            if not row['HBEN'] in location_sets['waterstandlocaties']['gdf']['LOC_ID'].values:
                error['missing_hben'] = True
                
            if not row['HBOVPS'] in location_sets['peilschalen']['gdf']['LOC_ID'].values:
                error['missing_hbovps'] = True
            
            if not row['HBENPS'] in location_sets['peilschalen']['gdf']['LOC_ID'].values:
                error['missing_hbenps'] = True
                
            if not row['PAR_ID'] in location_sets['hoofdlocaties']['gdf']['LOC_ID'].values:
                error['missing_hloc'] = True
            
            else:
                if not any([re.match(loc,loc_id) for loc in xy_ignore_df['internalLocation']]):
                    if not par_gdf[par_gdf['LOC_ID'] == row['PAR_ID']]['geometry'].values[0].equals(row['geometry']):
                        error['xy_not_same'] = True
                    
            if any(error.values()):        
                error['type'] = sub_type
                error['functie'] = loc_functie
        
        elif set_name == 'hoofdlocaties':
            if not re.match(f'[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*',loc_name):
                error['name_error'] = True        
                
        elif set_name == 'waterstandlocaties':
            if not re.match(f'[A-Z0-9 ]*_{caw_code}-w_.*',loc_name):
                error['name_error'] = True
            
            if not error['name_error']:
                caw_name = re.match(f'([A-Z0-9 ]*)_',loc_name).group(1)
                if not all(location_gdf[location_gdf['LOC_ID'].str.match(
                        f'..{caw_code}')]['LOC_NAME'].str.match(
                            f'({caw_name}_{caw_code}-w)')
                            ):
                                error['caw_name_inconsistent'] = True
    
            if not row['PEILSCHAAL'] in location_sets['peilschalen']['gdf']['LOC_ID'].values:
                error['missing_peilschaal'] = True
                
            if not loc_id in int_locs:
                error['missing_in_map'] = True
        
        if any(error.values()):
            loc_set_errors['locationId'].append(loc_id)
            loc_set_errors['caw_name'].append(caw_name)
            loc_set_errors['caw_code'].append(caw_code)
            loc_set_errors['csv_file'].append(csv_file)
            loc_set_errors['location_name'].append(loc_name)
            for key, value in error.items():
                loc_set_errors[key].append(value)
        
        # miss_locs = [loc for loc in int_locs if not loc in location_set['gdf']['LOC_ID'].values]
        # #for loc_id in miss_locs:
        # loc_set_errors['locationId'].append(miss_locs)
        # loc_set_errors['csv_file'].append([csv_file] * len(miss_locs))
        # loc_set_errors['location_name'].append([''] * len(miss_locs))
        # loc_set_errors['missing_in_set'].append([True] * len(miss_locs))
        # for key in ['loc_name_error','missing_in_map','missing_peilschaal']:
        #     loc_set_errors[key].append([False] * len(miss_locs))

consistency_df['locSet error'] = pd.DataFrame(loc_set_errors)
#opname in samenvatting

if len(consistency_df['locSet error']) == 0:
    logging.info('geen fouten in locationSets')
else:
    logging.warning('{} fouten in locationSets'.format(len(consistency_df['locSet error'])))        

#%% wegschrijven naar excel
inhoudsopgave = consistency_df['inhoudsopgave']
inhoudsopgave.index = inhoudsopgave['werkblad']
summary = {key:len(df) for key, df in consistency_df.items() if key in warning_sheets}
   
#lees input xlsx en gooi alles weg behalve de fixed_sheets
book = load_workbook(consistency_out_xlsx)
for worksheet in book.worksheets:
    if not worksheet.title in fixed_sheets:
        book.remove(worksheet)

# voeg samenvatting toe
worksheet = book.create_sheet('samenvatting',1)
worksheet.sheet_properties.tabColor = '92D050'
worksheet.append(['controle','aantal','beschrijving'])
for cell in worksheet['{}'.format(worksheet.max_row)]:
    cell.font = Font(bold=True)
    
for key, value in summary.items():
    worksheet.append([key,value,inhoudsopgave.loc[key]['beschrijving']])
    if value > 0:
       worksheet[worksheet.max_row][1].fill = PatternFill(fgColor='FF0000', fill_type='solid')
    else:
        worksheet[worksheet.max_row][1].fill = PatternFill(fgColor='92D050', fill_type='solid')

worksheet.column_dimensions['A'].width=40
worksheet.column_dimensions['C'].width = 100
worksheet.auto_filter.ref = worksheet.dimensions

xls_writer = pd.ExcelWriter(consistency_out_xlsx, engine='openpyxl')
xls_writer.book = book

for sheet_name, df in consistency_df.items():
        if (not sheet_name in fixed_sheets) & (not df.empty):
            if df.index.name == None:
                df.to_excel(xls_writer, sheet_name=sheet_name, index=False)
            else:
                df.to_excel(xls_writer, sheet_name=sheet_name, index=True)
            worksheet = xls_writer.sheets[sheet_name]
            for col in worksheet.columns:
                worksheet.column_dimensions[col[0].column_letter].width = 20
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.freeze_panes = worksheet['B2']
            if not df.empty:
                if (sheet_name in warning_sheets):
                    worksheet.sheet_properties.tabColor = 'FF0000'
                else:
                    worksheet.sheet_properties.tabColor = '92D050'
                    
xls_writer.book.active = xls_writer.book['samenvatting']

xls_writer.save()

#%% updaten csv's
def update_date(row,mpt_df,date_threshold):
    ''' wegschrijven van de start- en end-date'''
    int_loc = row['LOC_ID']
    if int_loc in mpt_df.index:
        start_date = mpt_df.loc[int_loc]['STARTDATE'].strftime('%Y%m%d')
        end_date = mpt_df.loc[int_loc]['ENDDATE']
        if end_date > date_threshold:
            end_date = pd.Timestamp(year=2100, month=1, day=1)
        end_date = end_date.strftime('%Y%m%d')
    else:
        start_date = row['START']
        end_date = row['EIND']
        
    return start_date, end_date

def update_histtag(row,grouper):
    ''' functie waarmee de laatste histag op aan waterstandsloc wordt toegekend '''
    return next((df.sort_values('total_max_end_dt', ascending=False)['serie'].values[0]
                 for loc_id, df 
                 in grouper 
                 if loc_id == row['LOC_ID']),None)


def update_peilschaal(row):
    ''' toekennen van de bovenstroomse en benedenstroomse peilschalen aan sublocaties '''
    result = {'HBOV':'','HBEN':''}
    
    for key in result.keys():
        df = waterstandloc_gdf.loc[waterstandloc_gdf['LOC_ID'] == row[key]]
        if not df.empty:
            result[key] = df['PEILSCHAAL'].values[0]
    
    return result['HBOV'], result['HBEN']

    

#def update_types(row):
    

date_threshold = mpt_df['ENDDATE'].max() - pd.Timedelta(weeks=26)

for locationSet, gdf in {'OPVLWATER_HOOFDLOC': hoofdloc_gdf,
                         'OPVLWATER_WATERSTANDEN_AUTO': waterstandloc_gdf,
                         'OPVLWATER_SUBLOC': subloc_gdf,}.items():
    logging.info(f'wegschrijven csv voor locationSet: {locationSet}')
    df = gdf.drop('geometry',axis=1)
    df[['START','EIND']] = df.apply(update_date, 
                                    args=(mpt_df, date_threshold), 
                                    axis=1,
                                    result_type="expand")
    
    if locationSet == 'OPVLWATER_WATERSTANDEN_AUTO':
        grouper = mpt_hist_tags_df.groupby(['fews_locid'])
        df['HIST_TAG'] = df.apply(update_histtag,
                                    args=[grouper],
                                    axis=1,
                                    result_type="expand")
        
    elif locationSet == 'OPVLWATER_SUBLOC':
        grouper = df.groupby(['PAR_ID'])
        par_types_df = grouper['TYPE'].unique().apply(lambda x: sorted(x)).transform(lambda x: '/'.join(x))
        df['PAR_ID'] = gdf['LOC_ID'].str[0:-1] + '0'
        df['ALLE_TYPES'] = df['PAR_ID'].apply(lambda x: par_types_df.loc[x])
        df[['HBOVPS','HBENPS']] = df.apply(update_peilschaal, 
                                    axis=1,
                                    result_type="expand")

    csv_file = csv_out.joinpath(fews_config.locationSets[locationSet]['csvFile']['file'])
    if csv_file.suffix == '':
        csv_file = Path(f'{csv_file}.csv')
    df.to_csv(csv_file, index=False)