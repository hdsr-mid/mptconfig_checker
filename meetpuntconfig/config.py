"""Module to read, check and write a HDSR meetpuntconfiguratie."""
__title__ = "histTags2mpt"
__description__ = "to evaluate a HDSR FEWS-config with a csv with CAW histTags"
__version__ = "0.1.0"
__author__ = "Daniel Tollenaar"
__author_email__ = "daniel@d2hydro.nl"
__license__ = "MIT License"

from meetpuntconfig.fews_utilities import FewsConfig, xml_to_dict
from pathlib import Path
import json
import numpy as np
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import sys
import re
from shapely.geometry import Point

pd.options.mode.chained_assignment = None


def idmap2tags(row, idmap):
    """Add FEWS-locationIds to hist_tags in df.apply() method."""
    exloc, expar = row["serie"].split("_", 1)
    fews_locs = [
        col["internalLocation"]
        for col in idmap
        if col["externalLocation"] == exloc and col["externalParameter"] == expar
    ]
    if len(fews_locs) == 0:
        fews_locs = np.NaN

    return fews_locs


def get_validation_attribs(validation_rules, int_pars=None, loc_type=None):
    """Get attributes from validationRules."""
    if int_pars is None:
        int_pars = [rule["parameter"] for rule in validation_rules]
    result = []
    for rule in validation_rules:
        if "type" in rule.keys():
            if rule["type"] == loc_type:
                if any(re.match(rule["parameter"], int_par) for int_par in int_pars):
                    for key, attribute in rule["extreme_values"].items():
                        if isinstance(attribute, list):
                            result += [value["attribute"] for value in attribute]
                        else:
                            result += [attribute]
        elif any(re.match(rule["parameter"], int_par) for int_par in int_pars):
            for key, attribute in rule["extreme_values"].items():
                if isinstance(attribute, list):
                    result += [value["attribute"] for value in attribute]
                else:
                    result += [attribute]
    return result


def update_hlocs(row, h_locs, mpt_df):
    """Add startdate and enddate op hoofdloc dataframe with df.apply() method."""
    loc_id = row.name
    start_date = row["STARTDATE"]
    end_date = row["ENDDATE"]

    if loc_id in h_locs:
        start_date = (
            mpt_df[mpt_df.index.str.contains(loc_id[0:-1])]["STARTDATE"].dropna().min()
        )

        end_date = (
            mpt_df[mpt_df.index.str.contains(loc_id[0:-1])]["ENDDATE"].dropna().max()
        )

    return start_date, end_date


def update_date(row, mpt_df, date_threshold):
    """Return start and end-date in df.apply() method."""
    int_loc = row["LOC_ID"]
    if int_loc in mpt_df.index:
        start_date = mpt_df.loc[int_loc]["STARTDATE"].strftime("%Y%m%d")
        end_date = mpt_df.loc[int_loc]["ENDDATE"]
        if end_date > date_threshold:
            end_date = pd.Timestamp(year=2100, month=1, day=1)
        end_date = end_date.strftime("%Y%m%d")
    else:
        start_date = row["START"]
        end_date = row["EIND"]

    return start_date, end_date


def update_histtag(row, grouper):
    """Assign last histTag to waterstandsloc in df.apply method."""
    return next(
        (
            df.sort_values("total_max_end_dt", ascending=False)["serie"].values[0]
            for loc_id, df in grouper
            if loc_id == row["LOC_ID"]
        ),
        None,
    )


def _sort_validation_attribs(rule):
    result = {}
    for key, value in rule.items():
        if isinstance(value, str):
            result[key] = [value]
        elif isinstance(value, list):
            periods = [val["period"] for val in value]
            attribs = [val["attribute"] for val in value]
            result[key] = [attrib for _, attrib in sorted(zip(periods, attribs))]
    return result


class MeetpuntConfig:
    """Meetpuntconfig class."""

    def __init__(self, config_path, log_level="INFO"):
        self.paths = dict()
        self.fews_config = None
        self.location_sets = dict()
        self.hist_tags = None
        self.hist_tags_ignore = None
        self.fixed_sheets = None
        self.idmap_files = None
        self.idmap_sections = None
        self.external_parameters_allowed = None
        self.consistency = None
        self.parameter_mapping = None
        self.validation_rules = None
        self.logging = logging
        self.hoofdloc = None
        self.subloc = None
        self.waterstandloc = None
        self.mswloc = None
        self.mpt_hist_tags = None
        self._locs_mapping = dict(
            hoofdlocaties="hoofdloc",
            sublocaties="subloc",
            waterstandlocaties="waterstandloc",
            mswlocaties="mswloc",
        )
        self.logging.basicConfig(level=os.environ.get("LOGLEVEL", log_level))

        self._read_config(Path(config_path))

    def _read_config(self, config_json):
        if config_json.exists():
            with open(config_json) as src:
                config = json.load(src)
                workdir = Path(config_json).parent
        else:
            self.logging.error(f"{config_json} does not exist")
            sys.exit()

        # add paths to config
        for key, path in config["paden"].items():
            path = Path(path)
            if not path.is_absolute():
                path = workdir.joinpath(path).resolve()
            if path.exists():
                self.paths[key] = path
            else:
                if path.suffix == "":
                    logging.warning(f"{path} does not exist. Folder will be created")
                    path.mkdir()
                else:
                    self.logging.error(
                        (
                            f"{path} does not exist. "
                            f"Please define existing file "
                            f"in {config_json}."
                        )
                    )
                    sys.exit()

        # add fews_config
        self.fews_config = FewsConfig(self.paths["fews_config"])

        # add location_sets
        for key, value in config["location_sets"].items():
            if value in self.fews_config.locationSets.keys():
                if "csvFile" in self.fews_config.locationSets[value].keys():
                    self.location_sets[key] = {
                        "id": value,
                        "gdf": self.fews_config.get_locations(value),
                    }
                else:
                    self.logging.error((f"{key} not a csvFile location-set"))
            else:
                self.logging.error(
                    (
                        f"locationSet {key} specified in {config_json} "
                        f"not in fews-config"
                    )
                )

        # add rest of config
        self.idmap_files = config["idmap_files"]
        self.idmap_sections = config["idmap_sections"]
        self.external_parameters_allowed = config["external_parameters_allowed"]
        self.parameter_mapping = config["parameter_mapping"]
        self.validation_rules = config["validation_rules"]
        self.fixed_sheets = config["fixed_sheets"]

        # read consistency df from input-excel
        self.consistency = pd.read_excel(
            self.paths["consistency_xlsx"], sheet_name=None, engine="openpyxl"
        )

        self.consistency = {
            key: value
            for key, value in self.consistency.items()
            if key in self.fixed_sheets
        }

    def _read_hist_tags(self, force=False):
        if (not self.hist_tags) or force:
            if "hist_tags_csv" in self.paths.keys():
                self.logging.info(f"reading histags: {self.paths['hist_tags_csv']}")
                dtype_cols = ["total_min_start_dt", "total_max_end_dt"]
                self.hist_tags = pd.read_csv(
                    self.paths["hist_tags_csv"],
                    parse_dates=dtype_cols,
                    sep=None,
                    engine="python",
                )

            for col in dtype_cols:
                if not pd.api.types.is_datetime64_dtype(self.hist_tags[col]):
                    self.logging.error(
                        (
                            f"col '{col}' in '{self.paths['hist_tags_csv']} "
                            "can't be converted to np.datetime64 format. "
                            "Check if values are dates."
                        )
                    )

                    sys.exit()

    def _read_hist_tags_ignore(self, force=False):
        if (not self.hist_tags_ignore) or force:
            if "mpt_ignore_csv" in self.paths.keys():
                self.logging.info(
                    f"Reading hist tags to be ingored from "
                    f"{self.paths['mpt_ignore_csv']}"
                )

                self.hist_tags_ignore = pd.read_csv(
                    self.paths["mpt_ignore_csv"], sep=None, header=0, engine="python"
                )

            elif "histTag_ignore" in self.consistency.keys():
                self.hist_tags_ignore = self.consistency["histTag_ignore"]
                self.logging.info(
                    f"Reading hist tags to be ignored from "
                    f"{self.paths['consistency_xlsx']}"
                )

            else:
                self.logging.error(
                    (
                        f"specify a histTag_ignore worksheet in "
                        f"{self.paths['consistency_xlsx']} or a csv-file "
                        "in the config-json"
                    )
                )

                sys.exit()
            self.hist_tags_ignore["UNKNOWN_SERIE"] = self.hist_tags_ignore[
                "UNKNOWN_SERIE"
            ].str.replace("#", "")

    def _get_idmaps(self, idmap_files=None):
        if not idmap_files:
            idmap_files = self.idmap_files
        idmaps = [
            xml_to_dict(self.fews_config.IdMapFiles[idmap])["idMap"]["map"]
            for idmap in idmap_files
        ]

        return [item for sublist in idmaps for item in sublist]

    def _read_locs(self):
        self.hoofdloc = self.fews_config.get_locations("OPVLWATER_HOOFDLOC")
        self.subloc = self.fews_config.get_locations("OPVLWATER_SUBLOC")
        self.waterstandloc = self.fews_config.get_locations(
            "OPVLWATER_WATERSTANDEN_AUTO"
        )

        self.mswloc = self.fews_config.get_locations("MSW_STATIONS")

    def _update_staff_gauge(self, row):
        """Assign upstream and downstream staff gauges to subloc."""
        result = {"HBOV": "", "HBEN": ""}

        for key in result.keys():
            df = self.waterstandloc.loc[self.waterstandloc["LOC_ID"] == row[key]]
            if not df.empty:
                result[key] = df["PEILSCHAAL"].values[0]

        return result["HBOV"], result["HBEN"]

    def hist_tags_to_mpt(self, sheet_name="mpt"):
        """Convert histTag-ids to mpt-ids."""
        if self.hist_tags is None:
            self._read_hist_tags()

        idmaps = self._get_idmaps()

        hist_tags_df = self.hist_tags.copy()
        hist_tags_df["fews_locid"] = hist_tags_df.apply(
            idmap2tags, args=[idmaps], axis=1
        )

        hist_tags_df = hist_tags_df[hist_tags_df["fews_locid"].notna()]
        mpt_hist_tags_df = hist_tags_df.explode("fews_locid").reset_index(drop=True)
        self.mpt_hist_tags = mpt_hist_tags_df
        mpt_df = pd.concat(
            [
                mpt_hist_tags_df.groupby(["fews_locid"], sort=False)[
                    "total_min_start_dt"
                ].min(),
                mpt_hist_tags_df.groupby(["fews_locid"], sort=False)[
                    "total_max_end_dt"
                ].max(),
            ],
            axis=1,
        )

        mpt_df = mpt_df.sort_index(axis=0)
        mpt_df.columns = ["STARTDATE", "ENDDATE"]
        mpt_df.index.name = "LOC_ID"
        kw_locs = list(mpt_df[mpt_df.index.str.contains("KW", regex=False)].index)
        h_locs = np.unique(["{}0".format(loc[0:-1]) for loc in kw_locs])
        h_locs_missing = [loc for loc in h_locs if loc not in list(mpt_df.index)]
        h_locs_df = pd.DataFrame(
            data={
                "LOC_ID": h_locs_missing,
                "STARTDATE": [pd.NaT] * len(h_locs_missing),
                "ENDDATE": [pd.NaT] * len(h_locs_missing),
            }
        )

        h_locs_df = h_locs_df.set_index("LOC_ID")
        mpt_df = pd.concat([mpt_df, h_locs_df], axis=0)
        mpt_df[["STARTDATE", "ENDDATE"]] = mpt_df.apply(
            update_hlocs, args=[h_locs, mpt_df], axis=1, result_type="expand"
        )
        mpt_df = mpt_df.sort_index()
        self.consistency["mpt"] = mpt_df

    def check_idmap_sections(self, sheet_name="idmap section error"):
        """Check if all KW/OW locations are in the correct section."""
        self.consistency[sheet_name] = pd.DataFrame(
            columns=[
                "bestand",
                "externalLocation",
                "externalParameter",
                "internalLocation",
                "internalParameter",
            ]
        )

        for idmap, subsecs in self.idmap_sections.items():
            for section_type, sections in subsecs.items():
                for section in sections:
                    if section_type == "KUNSTWERKEN":
                        prefix = "KW"
                    if section_type == "WATERSTANDLOCATIES":
                        prefix = "OW"
                    if section_type == "MSWLOCATIES":
                        prefix = "(OW|KW)"
                    pattern = fr"{prefix}\d{{6}}$"
                    idmapping = xml_to_dict(
                        self.fews_config.IdMapFiles[idmap], **section
                    )["idMap"]["map"]

                    idmap_wrong_section = [
                        idmap
                        for idmap in idmapping
                        if not bool(re.match(pattern, idmap["internalLocation"]))
                    ]
                    if idmap_wrong_section:
                        section_start = (
                            section["section_start"]
                            if "section_start" in section.keys()
                            else ""
                        )
                        section_end = (
                            section["section_end"]
                            if "section_end" in section.keys()
                            else ""
                        )
                        self.logging.warning(
                            (
                                f"{len(idmap_wrong_section)} "
                                f"internalLocations not {prefix}XXXXXX "
                                f"between {section_start} and {section_end} "
                                f"in {idmap}."
                            )
                        )

                        df = pd.DataFrame(idmap_wrong_section)
                        df["sectie"] = section_start
                        df["bestand"] = idmap
                        self.consistency[sheet_name] = pd.concat(
                            [self.consistency[sheet_name], df], axis=0
                        )

    def check_missing_hist_tags(self, sheet_name="histTags noMatch"):
        """Check if hisTags are missing in config."""
        if self.hist_tags_ignore is None:
            self._read_hist_tags_ignore()

        if self.hist_tags is None:
            self._read_hist_tags()

        hist_tags_df = self.hist_tags.copy()

        idmaps = self._get_idmaps()

        hist_tags_df["fews_locid"] = self.hist_tags.apply(
            idmap2tags, args=[idmaps], axis=1
        )

        hist_tags_no_match_df = hist_tags_df[hist_tags_df["fews_locid"].isna()]
        hist_tags_no_match_df = hist_tags_no_match_df[
            ~hist_tags_no_match_df["serie"].isin(self.hist_tags_ignore["UNKNOWN_SERIE"])
        ]

        hist_tags_no_match_df = hist_tags_no_match_df.drop("fews_locid", axis=1)
        hist_tags_no_match_df.columns = ["UNKNOWN_SERIE", "STARTDATE", "ENDDATE"]
        hist_tags_no_match_df = hist_tags_no_match_df.set_index("UNKNOWN_SERIE")
        self.consistency[sheet_name] = hist_tags_no_match_df

        if not self.consistency[sheet_name].empty:
            self.logging.warning(
                "{} histTags not in idMaps".format(len(self.consistency[sheet_name]))
            )

        else:
            self.logging.info("all histTags in idMaps")

    def check_ignored_hist_tags(
        self, sheet_name="histTags ignore match", idmap_files=["IdOPVLWATER"]
    ):
        """Check if ignored histTags do match with idmap."""
        if self.hist_tags_ignore is None:
            self._read_hist_tags_ignore()
        if self.hist_tags is None:
            self._read_hist_tags()
        hist_tags_opvlwater_df = self.hist_tags.copy()

        idmaps = self._get_idmaps(idmap_files=idmap_files)

        hist_tags_opvlwater_df["fews_locid"] = self.hist_tags.apply(
            idmap2tags, args=[idmaps], axis=1
        )

        hist_tags_opvlwater_df = hist_tags_opvlwater_df[
            hist_tags_opvlwater_df["fews_locid"].notna()
        ]

        hist_tag_ignore_match_df = self.hist_tags_ignore[
            self.hist_tags_ignore["UNKNOWN_SERIE"].isin(hist_tags_opvlwater_df["serie"])
        ]

        hist_tag_ignore_match_df = hist_tag_ignore_match_df.set_index("UNKNOWN_SERIE")
        self.consistency[sheet_name] = hist_tag_ignore_match_df

        if not self.consistency[sheet_name].empty:
            self.logging.warning(
                (
                    f"{len(self.consistency[sheet_name])} "
                    r"histTags should not be in histTags ignore."
                )
            )

        else:
            self.logging.info("hisTags ignore list consistent with idmaps")

    def check_double_idmaps(self, sheet_name="idmaps double"):
        """Check if identical idmaps are doubled."""
        self.consistency[sheet_name] = pd.DataFrame(
            columns=[
                "bestand",
                "externalLocation",
                "externalParameter",
                "internalLocation",
                "internalParameter",
            ]
        )
        for idmap_file in self.idmap_files:
            idmaps = self._get_idmaps(idmap_files=[idmap_file])

            idmap_doubles = [idmap for idmap in idmaps if idmaps.count(idmap) > 1]
            if len(idmap_doubles) > 0:
                idmap_doubles = list(
                    {
                        idmap["externalLocation"]: idmap for idmap in idmap_doubles
                    }.values()
                )
                df = pd.DataFrame(
                    idmap_doubles,
                    columns=[
                        "internalLocation",
                        "externalLocation",
                        "internalParameter",
                        "externalParameter",
                    ],
                )

                df["bestand"] = idmap_file
                self.consistency[sheet_name] = pd.concat(
                    [self.consistency[sheet_name], df], axis=0
                )
                self.logging.warning(
                    "{} double idmap(s) in {}".format(len(idmap_doubles), idmap_file)
                )
            else:
                self.logging.info("No double idmaps in {}".format(idmap_file))

    def check_missing_pars(self, sheet_name="pars missing"):
        """Check if internal parameters in idmaps are missing in paramters.xml."""
        config_parameters = list(
            self.fews_config.get_parameters(dict_keys="parameters").keys()
        )

        idmaps = self._get_idmaps()
        id_map_parameters = [id_map["internalParameter"] for id_map in idmaps]
        params_missing = [
            parameter
            for parameter in id_map_parameters
            if parameter not in config_parameters
        ]

        if len(params_missing) == 0:
            self.logging.info("all internal paramters are in config")
        else:
            self.logging.warning(
                "{} parameter(s) in idMaps are missing in config".format(
                    len(params_missing)
                )
            )

            self.consistency[sheet_name] = pd.DataFrame({"parameters": params_missing})

    #            self.consistency[sheet_name] = self.consistency[sheet_name].set_index(
    #               "parameters"
    #            )

    def check_hloc_consistency(self, sheet_name="hloc error"):
        """Check if all sublocs of same hloc have consistent parameters."""
        if "xy_ignore" in self.consistency.keys():
            xy_ignore_df = self.consistency["xy_ignore"]
        else:
            xy_ignore_df = pd.DataFrame({"internalLocation": [], "x": [], "y": []})

        if self.hoofdloc is None:
            self._read_locs()

        hloc_errors = {
            "LOC_ID": [],
            "SUB_LOCS": [],
            "LOC_NAME": [],
            "GEOMETRY": [],
            "SYSTEEM": [],
            "RAYON": [],
            "KOMPAS": [],
        }

        grouper = self.subloc.groupby("PAR_ID")
        par_dict = {
            "LOC_ID": [],
            "LOC_NAME": [],
            "X": [],
            "Y": [],
            "ALLE_TYPES": [],
            "START": [],
            "EIND": [],
            "SYSTEEM": [],
            "RAYON": [],
            "KOMPAS": [],
        }

        for loc_id, gdf in grouper:
            caw_code = loc_id[2:-2]
            errors = dict.fromkeys(
                ["LOC_NAME", "GEOMETRY", "SYSTEEM", "RAYON", "KOMPAS"], False
            )
            fields = dict.fromkeys(par_dict.keys(), None)
            fields["LOC_ID"] = loc_id

            loc_names = np.unique(
                gdf["LOC_NAME"]
                .str.extract(pat=f"([A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*)")
                .values
            )

            if not len(loc_names) == 1:
                errors["LOC_NAME"] = ",".join(loc_names)
            else:
                fields["LOC_NAME"] = loc_names[0]

            if any([re.match(loc, loc_id) for loc in xy_ignore_df["internalLocation"]]):
                fields["X"], fields["Y"] = next(
                    [row["x"], row["y"]]
                    for index, row in xy_ignore_df.iterrows()
                    if re.match(row["internalLocation"], loc_id)
                )

            else:
                geoms = gdf["geometry"].unique()
                if not len(geoms) == 1:
                    errors["GEOMETRY"] = ",".join(
                        [f"({geom.x} {geom.y})" for geom in geoms]
                    )

                else:
                    fields["X"] = geoms[0].x
                    fields["Y"] = geoms[0].y

            all_types = list(gdf["TYPE"].unique())
            all_types.sort()
            fields["ALLE_TYPES"] = "/".join(all_types)
            fields["START"] = gdf["START"].min()
            fields["EIND"] = gdf["EIND"].max()
            for attribuut in ["SYSTEEM", "RAYON", "KOMPAS"]:
                vals = gdf[attribuut].unique()
                if not len(vals) == 1:
                    errors[attribuut] = ",".join(vals)
                else:
                    fields[attribuut] = vals[0]
            if None not in fields.values():
                for key, value in fields.items():
                    par_dict[key].append(value)

            if any(errors.values()):
                hloc_errors["LOC_ID"].append(loc_id)
                hloc_errors["SUB_LOCS"].append(",".join(gdf["LOC_ID"].values))
                for key, value in errors.items():
                    if value is False:
                        value = ""
                    hloc_errors[key].append(value)
        self.consistency[sheet_name] = pd.DataFrame(hloc_errors)
        if self.consistency[sheet_name].empty:
            self.logging.info("no consistency errors. Hlocs rewritten from sublocs")
            par_gdf = pd.DataFrame(par_dict)
            columns = list(self.hoofdloc.columns)
            drop_cols = [
                col
                for col in self.hoofdloc.columns
                if (col in par_gdf.columns) & (not col == "LOC_ID")
            ]

            drop_cols = drop_cols + ["geometry"]
            self.hoofdloc = self.hoofdloc.drop(drop_cols, axis=1)
            self.hoofdloc = par_gdf.merge(self.hoofdloc, on="LOC_ID")
            self.hoofdloc["geometry"] = self.hoofdloc.apply(
                (lambda x: Point(float(x["X"]), float(x["Y"]))), axis=1
            )
            self.hoofdloc = self.hoofdloc[columns]
        else:
            self.logging.warning(
                "{} Errors in consistency hlocs".format(
                    len(self.consistency[sheet_name])
                )
            )
            self.logging.warning(
                (
                    "Hoofdlocaties will only be re-written "
                    "when consistency errors are resolved"
                )
            )

    def check_expar_errors_intloc_missing(
        self, expar_sheet="exPar error", intloc_sheet="intLoc missing"
    ):
        """Check on wrong external parameters and missing internal locations."""
        expars_allowed = self.external_parameters_allowed

        if self.hoofdloc is None:
            self._read_locs()

        ex_par_errors = {
            "internalLocation": [],
            "locationType": [],
            "exParError": [],
            "types": [],
            "FQ": [],
            "I.X": [],
            "IX.": [],
            "SS./SM.": [],
        }
        int_loc_missing = []
        idmap_df = pd.DataFrame.from_dict(self._get_idmaps(["IdOPVLWATER"]))
        for int_loc, loc_group in idmap_df.groupby("internalLocation"):
            errors = dict.fromkeys(["I.X", "IX.", "FQ", "SS./SM."], False)
            ex_pars = np.unique(loc_group["externalParameter"].values)
            ex_pars_gen = [re.sub(r"\d", ".", ex_par) for ex_par in ex_pars]
            if int_loc in self.hoofdloc["LOC_ID"].values:
                loc_properties = self.hoofdloc[self.hoofdloc["LOC_ID"] == int_loc]
                loc_type = "hoofdloc"
            elif int_loc in self.subloc["LOC_ID"].values:
                loc_properties = self.subloc[self.subloc["LOC_ID"] == int_loc]
                loc_type = "subloc"
                regexes = ["HR.$"]
            elif int_loc in self.waterstandloc["LOC_ID"].values:
                loc_type = "waterstandloc"
            elif int_loc in self.mswloc["LOC_ID"].values:
                loc_type = "mswloc"
            else:
                loc_type = None
                int_loc_missing += [int_loc]
            if loc_type in ["hoofdloc", "subloc"]:
                all_types = loc_properties["ALLE_TYPES"].values[0].split("/")
                all_types = [item.lower() for item in all_types]
            elif loc_type == "waterstandloc":
                all_types = ["waterstandloc"]
            if loc_type == "subloc":
                sub_type = self.subloc[self.subloc["LOC_ID"] == int_loc]["TYPE"].values[
                    0
                ]

                regexes += [
                    j
                    for i in [
                        values
                        for keys, values in expars_allowed.items()
                        if keys in all_types
                    ]
                    for j in i
                ]

                regexes += list(dict.fromkeys(regexes))
                ex_par_error = [
                    ex_par
                    for ex_par in ex_pars
                    if not any(
                        [
                            regex.match(ex_par)
                            for regex in [re.compile(rex) for rex in regexes]
                        ]
                    )
                ]

                if sub_type == "schuif":
                    if not any(
                        [ex_par for ex_par in ex_pars_gen if ex_par in ["SS.", "SM."]]
                    ):
                        errors["SS./SM."] = True

                if any(
                    [
                        ex_par
                        for ex_par in ex_pars_gen
                        if ex_par in ["I.B", "I.H", "I.L"]
                    ]
                ):
                    if not any(
                        [
                            ex_par
                            for ex_par in ex_pars_gen
                            if ex_par in ["IB.", "IH.", "IL."]
                        ]
                    ):
                        errors["IX."] = True

                elif any(
                    [
                        ex_par
                        for ex_par in ex_pars_gen
                        if ex_par in ["IB.", "IH.", "IL."]
                    ]
                ):
                    errors["I.X"] = True

                if "FQ." in ex_pars_gen:
                    if not any(
                        [
                            ex_par
                            for ex_par in ex_pars_gen
                            if ex_par in ["IB.", "IH.", "IL.", "I.B", "I.H", "I.L"]
                        ]
                    ):
                        errors["FQ"] = True

            elif loc_type == "hoofdloc":
                regexes = ["HS.$", "QR.$", "QS.$", "WR", "WS"]
                ex_par_error = [
                    ex_par
                    for ex_par in ex_pars
                    if not any(
                        [
                            regex.match(ex_par)
                            for regex in [re.compile(rex) for rex in regexes]
                        ]
                    )
                ]

            else:
                ex_par_error = []

            if len(ex_par_error) > 0 | any(errors.values()):
                ex_par_errors["internalLocation"].append(int_loc)
                ex_par_errors["locationType"].append(loc_type)
                ex_par_errors["exParError"].append(",".join(ex_par_error))
                ex_par_errors["types"].append(",".join(all_types))
                for key, value in errors.items():
                    ex_par_errors[key].append(value)

        self.consistency[expar_sheet] = pd.DataFrame(ex_par_errors)
        self.consistency[intloc_sheet] = pd.DataFrame(
            {"internalLocation": int_loc_missing}
        )

        if len(self.consistency[expar_sheet]) == 0:
            self.logging.info("geen ExPar errors")
        else:
            self.logging.warning(
                "{} locaties met ExPar errors".format(
                    len(self.consistency[expar_sheet])
                )
            )

        if len(self.consistency[intloc_sheet]) == 0:
            self.logging.info("All internal locations are in locationSets")
        else:
            self.logging.warning(
                "{} Internal locations are not in locationSets.".format(
                    len(self.consistency[intloc_sheet])
                )
            )

    def check_expar_missing(self, sheet_name="exPar missing"):
        """Check if external paramters are missing on locations."""
        ex_par_missing = {
            "internalLocation": [],
            "exPars": [],
            "QR": [],
            "QS": [],
            "HS": [],
        }
        if self.hoofdloc is None:
            self._read_locs()
        idmap_df = pd.DataFrame.from_dict(self._get_idmaps(["IdOPVLWATER"]))
        for index, row in self.hoofdloc.iterrows():
            missings = dict.fromkeys(["QR", "QS", "HS"], False)
            int_loc = row["LOC_ID"]

            loc_group = next(
                (
                    df
                    for loc, df in idmap_df.groupby("internalLocation")
                    if loc == int_loc
                ),
                pd.DataFrame(),
            )

            if not loc_group.empty:
                ex_pars = np.unique(loc_group["externalParameter"].values)
                ex_pars_gen = [re.sub(r"\d", ".", ex_par) for ex_par in ex_pars]
            else:
                ex_pars = []
                ex_pars_gen = []
            if not ("HS." in ex_pars_gen):
                missings["HS"] = True
            if not ("QR." in ex_pars_gen):
                missings["QR"] = True
            if not ("QS." in ex_pars_gen):
                missings["QS"] = True
            if any(missings.values()):
                ex_par_missing["internalLocation"].append(int_loc)
                ex_par_missing["exPars"].append(",".join(ex_pars))
                for key, value in missings.items():
                    ex_par_missing[key].append(value)

        self.consistency[sheet_name] = pd.DataFrame(ex_par_missing)

        if len(self.consistency[sheet_name]) == 0:
            self.logging.info("No ExPar missing")
        else:
            self.logging.warning(
                "{} Locations with ExPar missing".format(
                    len(self.consistency[sheet_name])
                )
            )

    def check_exloc_intloc_consistency(self, sheet_name="exLoc error"):
        """Check if external locations are consistent with internal locations."""
        ex_loc_errors = {"internalLocation": [], "externalLocation": []}

        idmap_df = pd.DataFrame.from_dict(self._get_idmaps(["IdOPVLWATER"]))
        for loc_group in idmap_df.groupby("externalLocation"):
            int_loc_error = []
            ex_loc = loc_group[0]
            int_locs = np.unique(loc_group[1]["internalLocation"].values)
            if len(ex_loc) == 3:
                if not bool(re.match("8..$", ex_loc)):
                    int_loc_error = [
                        int_loc
                        for int_loc in int_locs
                        if not bool(re.match(f"...{ex_loc}..$", int_loc))
                    ]
                else:
                    for loc_type in ["KW", "OW"]:
                        int_locs_select = [
                            int_loc
                            for int_loc in int_locs
                            if bool(re.match(f"{loc_type}.", int_loc))
                        ]
                        if (
                            len(
                                np.unique([int_loc[:-1] for int_loc in int_locs_select])
                            )
                            > 1
                        ):
                            int_loc_error += list(int_locs_select)
            if len(ex_loc) == 4:
                if not bool(re.match(".8..$", ex_loc)):
                    int_loc_error += [
                        int_loc
                        for int_loc in int_locs
                        if not bool(re.match(f"..{ex_loc}..$", int_loc))
                    ]
                else:
                    for loc_type in ["KW", "OW"]:
                        int_locs_select = [
                            int_loc
                            for int_loc in int_locs
                            if bool(re.match(f"{loc_type}.", int_loc))
                        ]
                        if (
                            len(
                                np.unique([int_loc[:-1] for int_loc in int_locs_select])
                            )
                            > 1
                        ):
                            int_loc_error += list(int_locs_select)
            if "exLoc_ignore" in self.consistency.keys():
                if (
                    int(ex_loc)
                    in self.consistency["exLoc_ignore"]["externalLocation"].values
                ):
                    int_loc_error = [
                        int_loc
                        for int_loc in int_loc_error
                        if int_loc
                        not in self.consistency["exLoc_ignore"][
                            self.consistency["exLoc_ignore"]["externalLocation"]
                            == int(ex_loc)
                        ]["internalLocation"].values
                    ]

            for int_loc in int_loc_error:
                ex_loc_errors["internalLocation"].append(int_loc)
                ex_loc_errors["externalLocation"].append(ex_loc)

        self.consistency[sheet_name] = pd.DataFrame(ex_loc_errors)

        if len(self.consistency[sheet_name]) == 0:
            self.logging.info("all external and internal locations consistent")
        else:
            self.logging.warning(
                "{} external locations inconsistent with internal locations".format(
                    len(self.consistency[sheet_name])
                )
            )

    def check_timeseries_logic(self, sheet_name="timeSeries error"):
        """Check if timeseries are consistent with internal locations and parameters."""
        if "TS800_ignore" in self.consistency.keys():
            ts_ignore_df = self.consistency["TS800_ignore"]
        else:
            ts_ignore_df = pd.DataFrame(
                {"internalLocation": [], "externalLocation": []}
            )

        idmap_df = pd.DataFrame.from_dict(self._get_idmaps(["IdOPVLWATER"]))

        if self.subloc is None:
            self._read_locs()
        idmap_subloc_df = idmap_df[
            idmap_df["internalLocation"].isin(self.subloc["LOC_ID"].values)
        ]

        idmap_subloc_df.loc[:, "type"] = idmap_subloc_df["internalLocation"].apply(
            (lambda x: self.subloc[self.subloc["LOC_ID"] == x]["TYPE"].values[0])
        )

        idmap_subloc_df.loc[:, "loc_groep"] = idmap_subloc_df["internalLocation"].apply(
            (lambda x: x[0:-1])
        )

        ts_errors = {
            "internalLocation": [],
            "eind": [],
            "internalParameters": [],
            "externalParameters": [],
            "externalLocations": [],
            "type": [],
            "fout": [],
        }

        for loc_group, group_df in idmap_subloc_df.groupby("loc_groep"):
            ex_locs = np.unique(group_df["externalLocation"].values)
            ex_locs_dict = {ex_loc: idx for idx, ex_loc in enumerate(ex_locs)}
            split_ts = [
                key
                for key in ex_locs_dict.keys()
                if any(
                    [
                        regex.match(key)
                        for regex in [re.compile(rex) for rex in ["8..", ".8.."]]
                    ]
                )
            ]

            ex_locs_skip = ts_ignore_df[
                ts_ignore_df["internalLocation"].isin(group_df["internalLocation"])
            ]["externalLocation"]

            split_ts = [
                key
                for key in split_ts
                if not str(key) in ex_locs_skip.values.astype(np.str)
            ]

            ex_locs_dict = {
                k: (
                    ex_locs_dict[k[1:]]
                    if (k[1:] in ex_locs_dict.keys()) and (k not in split_ts)
                    else v
                )
                for (k, v) in ex_locs_dict.items()
            }

            org_uniques = np.unique(
                [val for key, val in ex_locs_dict.items() if key not in split_ts]
            )
            if (len(org_uniques) == 1) & (len(split_ts) == 1):
                ex_locs_dict = {
                    k: (org_uniques[0] if k in split_ts else v)
                    for (k, v) in ex_locs_dict.items()
                }

            group_df["ex_loc_group"] = group_df["externalLocation"].apply(
                (lambda x: ex_locs_dict[x])
            )

            for int_loc, loc_df in group_df.groupby("internalLocation"):
                sub_type = self.subloc[self.subloc["LOC_ID"] == int_loc]["TYPE"].values[
                    0
                ]

                end_time = pd.to_datetime(
                    self.subloc[self.subloc["LOC_ID"] == int_loc]["EIND"].values[0]
                )

                ex_pars = np.unique(loc_df["externalParameter"].values)
                int_pars = np.unique(loc_df["internalParameter"].values)
                ex_locs = np.unique(loc_df["externalLocation"].values)
                if sub_type in ["krooshek", "debietmeter"]:
                    if any([re.match("HR.", ex_par) for ex_par in ex_pars]):
                        ts_errors["internalLocation"].append(int_loc)
                        ts_errors["eind"].append(end_time)
                        ts_errors["internalParameters"].append(",".join(int_pars))
                        ts_errors["externalParameters"].append(",".join(ex_pars))
                        ts_errors["externalLocations"].append(",".join(ex_locs))
                        ts_errors["type"].append(sub_type)
                        ts_errors["fout"].append(f"{sub_type} met stuurpeil")

                else:
                    if not any([re.match("HR.", ex_par) for ex_par in ex_pars]):
                        if any(
                            [
                                re.match("HR.", ex_par)
                                for ex_par in np.unique(group_df["externalParameter"])
                            ]
                        ):
                            if sub_type not in ["totaal", "vispassage"]:
                                if pd.Timestamp.now() < end_time:
                                    sp_locs = np.unique(
                                        group_df[
                                            group_df["externalParameter"].str.match(
                                                "HR."
                                            )
                                        ]["internalLocation"]
                                    )
                                    ts_errors["internalLocation"].append(int_loc)
                                    ts_errors["eind"].append(end_time)
                                    ts_errors["internalParameters"].append(
                                        ",".join(int_pars)
                                    )

                                    ts_errors["externalParameters"].append(
                                        ",".join(ex_pars)
                                    )

                                    ts_errors["externalLocations"].append(
                                        ",".join(ex_locs)
                                    )

                                    ts_errors["type"].append(sub_type)
                                    ts_errors["fout"].append(
                                        (
                                            f"{sub_type} zonder stuurpeil "
                                            f"({','.join(sp_locs)} wel)"
                                        )
                                    )
                    else:
                        time_series = loc_df.groupby(
                            ["ex_loc_group", "externalParameter"]
                        )
                        sp_series = [
                            series
                            for series in time_series
                            if bool(re.match("HR.", series[0][1]))
                        ]
                        for idx, series in enumerate(sp_series):
                            ex_par = series[0][1]
                            ex_locs = series[1]["externalLocation"]
                            int_par = np.unique(series[1]["internalParameter"])
                            if len(int_par) > 1:
                                ts_errors["internalLocation"].append(int_loc)
                                ts_errors["eind"].append(end_time)
                                ts_errors["internalParameters"].append(
                                    ",".join(int_pars)
                                )

                                ts_errors["externalParameters"].append(
                                    ",".join(ex_pars)
                                )

                                ts_errors["externalLocations"].append(",".join(ex_locs))
                                ts_errors["type"].append(sub_type)
                                ts_errors["fout"].append(
                                    (
                                        f'{",".join(int_par)} coupled to 1 sp-series ('
                                        f"exPar: {ex_par}, exLoc(s)): "
                                        f'{",".join(ex_locs)}'
                                    )
                                )
                            other_series = [
                                series
                                for idy, series in enumerate(sp_series)
                                if not idy == idx
                            ]

                            other_int_pars = [
                                np.unique(series[1]["internalParameter"])
                                for series in other_series
                            ]

                            if len(other_int_pars) > 0:
                                other_int_pars = np.concatenate(other_int_pars)

                            conflicting_pars = [
                                par for par in int_par if par in other_int_pars
                            ]

                            if len(conflicting_pars) > 0:
                                # 2 sp series gekoppeld aan dezelfde fews parameter
                                ts_errors["internalLocation"].append(int_loc)
                                ts_errors["eind"].append(end_time)
                                ts_errors["internalParameters"].append(
                                    ",".join(int_pars)
                                )

                                ts_errors["externalParameters"].append(
                                    ",".join(ex_pars)
                                )

                                ts_errors["externalLocations"].append(",".join(ex_locs))
                                ts_errors["type"].append(sub_type)
                                ts_errors["fout"].append(
                                    (
                                        f'{",".join(conflicting_pars)} gekoppeld aan '
                                        f"sp-serie (exPar: {ex_par}, exLoc(s)):"
                                        f'{",".join(ex_locs)}'
                                    )
                                )

        self.consistency[sheet_name] = pd.DataFrame(ts_errors)
        if len(self.consistency[sheet_name]) == 0:
            self.logging.info(
                (
                    "logical coupling of all timeseries to internal "
                    "locations/parameters"
                )
            )
        else:
            self.logging.warning(
                (
                    f"{len(self.consistency[sheet_name])} timeseries "
                    r"coupled illogical to internal locations/parameters"
                )
            )

    def check_validation_rules(self, sheet_name="validation error"):
        """Check if validation rules are consistent."""
        valid_errors = {
            "internalLocation": [],
            "start": [],
            "eind": [],
            "internalParameters": [],
            "fout_type": [],
            "fout_beschrijving": [],
        }

        location_sets_dict = xml_to_dict(
            self.fews_config.RegionConfigFiles["LocationSets"]
        )["locationSets"]["locationSet"]

        if self.hoofdloc is None:
            self._read_locs()

        for set_name in self.validation_rules.keys():
            location_set_meta = next(
                loc_set
                for loc_set in location_sets_dict
                if loc_set["id"] == self.location_sets[set_name]["id"]
            )["csvFile"]

            location_set_gdf = getattr(self, self._locs_mapping[set_name])
            attrib_files = location_set_meta["attributeFile"]
            if not isinstance(attrib_files, list):
                attrib_files = [attrib_files]

            attrib_files = [
                attrib_file
                for attrib_file in attrib_files
                if "attribute" in attrib_file.keys()
            ]

            for attrib_file in attrib_files:
                attribs = attrib_file["attribute"]
                join_id = attrib_file["id"].replace("%", "")
                if not isinstance(attrib_file["attribute"], list):
                    attribs = [attribs]
                attribs = [
                    attrib["number"].replace("%", "")
                    for attrib in attribs
                    if "number" in attrib.keys()
                ]

                attrib_df = pd.read_csv(
                    self.fews_config.MapLayerFiles[
                        attrib_file["csvFile"].replace(".csv", "")
                    ],
                    sep=None,
                    engine="python",
                )

                attrib_df.rename(columns={join_id: "LOC_ID"}, inplace=True)
                drop_cols = [
                    col for col in attrib_df if col not in attribs + ["LOC_ID"]
                ]

                attrib_df = attrib_df.drop(columns=drop_cols, axis=1)
                location_set_gdf = location_set_gdf.merge(
                    attrib_df, on="LOC_ID", how="outer"
                )

            validation_rules = self.validation_rules[set_name]
            validaton_attributes = get_validation_attribs(validation_rules)
            idmap_df = pd.DataFrame.from_dict(self._get_idmaps(["IdOPVLWATER"]))
            params_df = pd.DataFrame.from_dict(
                {
                    int_loc: [df["internalParameter"].values]
                    for int_loc, df in idmap_df.groupby("internalLocation")
                },
                orient="index",
                columns=["internalParameters"],
            )

            for (idx, row) in location_set_gdf.iterrows():
                int_loc = row["LOC_ID"]
                row = row.dropna()
                # if set_name == 'sublocaties':
                #     loc_type = row['TYPE']
                if int_loc in params_df.index:
                    int_pars = np.unique(params_df.loc[int_loc]["internalParameters"])
                else:
                    int_pars = []

                attribs_required = get_validation_attribs(validation_rules, int_pars)
                attribs_missing = [
                    attrib for attrib in attribs_required if attrib not in row.keys()
                ]

                attribs_obsolete = [
                    attrib
                    for attrib in validaton_attributes
                    if (attrib not in attribs_required) and (attrib in row.keys())
                ]

                attribs = [
                    attrib
                    for attrib in attribs_required
                    if attrib not in attribs_missing
                ]

                for key, value in {
                    "missend": attribs_missing,
                    "overbodig": attribs_obsolete,
                }.items():
                    if len(value) > 0:
                        valid_errors["internalLocation"] += [int_loc]
                        valid_errors["start"] += [row["START"]]
                        valid_errors["eind"] += [row["EIND"]]
                        valid_errors["internalParameters"] += [",".join(int_pars)]
                        valid_errors["fout_type"] += [key]
                        valid_errors["fout_beschrijving"] += [",".join(value)]
                for validation_rule in validation_rules:
                    errors = {"fout_type": None, "fout_beschrijving": []}
                    param = validation_rule["parameter"]
                    if any(re.match(param, int_par) for int_par in int_pars):
                        rule = validation_rule["extreme_values"]
                        rule = _sort_validation_attribs(rule)
                        if all(key in ["hmax", "hmin"] for key in rule.keys()):
                            for hmin, hmax in zip(rule["hmin"], rule["hmax"]):
                                if all(attrib in row.keys() for attrib in [hmin, hmax]):
                                    if row[hmax] < row[hmin]:
                                        errors["fout_type"] = "waarde"
                                        errors["fout_beschrijving"] += [
                                            f"{hmax} < {hmin}"
                                        ]

                        elif all(
                            key in rule.keys()
                            for key in ["hmax", "smax", "smin", "hmin"]
                        ):
                            hmax = rule["hmax"][0]
                            hmin = rule["hmin"][0]
                            for smin, smax in zip(rule["smin"], rule["smax"]):
                                if all(attrib in row.keys() for attrib in [smin, smax]):
                                    if row[smax] <= row[smin]:
                                        errors["fout_type"] = "waarde"
                                        errors["fout_beschrijving"] += [
                                            f"{smax} <= {smin}"
                                        ]

                                    if row[hmax] < row[smax]:
                                        errors["fout_type"] = "waarde"
                                        errors["fout_beschrijving"] += [
                                            f"{'hmax'} < {smax}"
                                        ]

                                    if row[smin] < row[hmin]:
                                        errors["fout_type"] = "waarde"
                                        errors["fout_beschrijving"] += [
                                            f"{smin} < {hmin}"
                                        ]

                    valid_errors["internalLocation"] += [row["LOC_ID"]] * len(
                        errors["fout_beschrijving"]
                    )

                    valid_errors["start"] += [row["START"]] * len(
                        errors["fout_beschrijving"]
                    )

                    valid_errors["eind"] += [row["EIND"]] * len(
                        errors["fout_beschrijving"]
                    )

                    valid_errors["internalParameters"] += [",".join(int_pars)] * len(
                        errors["fout_beschrijving"]
                    )

                    valid_errors["fout_type"] += [errors["fout_type"]] * len(
                        errors["fout_beschrijving"]
                    )

                    valid_errors["fout_beschrijving"] += errors["fout_beschrijving"]

        self.consistency[sheet_name] = pd.DataFrame(valid_errors)
        self.consistency[sheet_name] = self.consistency[sheet_name].drop_duplicates()

        if len(self.consistency[sheet_name]) == 0:
            self.logging.info("No missing incorrect validation rules")
        else:
            self.logging.warning(
                "{} validation rules contain errors/are missing".format(
                    len(self.consistency[sheet_name])
                )
            )

    def check_intpar_expar_consistency(self, sheet_name="par mismatch"):
        """Check if internal and external parameters are consistent."""
        par_errors = {
            "internalLocation": [],
            "internalParameter": [],
            "externalParameter": [],
            "fout": [],
        }

        # internal_parameters = [mapping[
        #     'internal'] for mapping in self.parameter_mapping]

        idmap_df = pd.DataFrame.from_dict(self._get_idmaps(["IdOPVLWATER"]))
        for idx, row in idmap_df.iterrows():
            error = None
            ext_par = None
            ext_par = [
                mapping["external"]
                for mapping in self.parameter_mapping
                if re.match(f'{mapping["internal"]}[0-9]', row["internalParameter"])
            ]

            if ext_par:
                if not any(re.match(par, row["externalParameter"]) for par in ext_par):
                    error = "parameter mismatch"
            else:
                error = "pars niet opgenomen in config"
            if error:
                par_errors["internalLocation"].append(row["internalLocation"])
                par_errors["internalParameter"].append(row["internalParameter"])
                par_errors["externalParameter"].append(row["externalParameter"])
                par_errors["fout"].append(error)

        self.consistency[sheet_name] = pd.DataFrame(par_errors)

        if len(self.consistency[sheet_name]) == 0:
            self.logging.info("geen regex fouten op interne en externe parameters")
        else:
            self.logging.warning(
                "{} regex fouten op interne en externe parameters".format(
                    len(self.consistency[sheet_name])
                )
            )

    def check_location_set_errors(self, sheet_name="locSet error"):
        """Check on errors in locationsets."""
        # fews_config = self.fews_config
        # config = self
        xy_ignore_df = self.consistency["xy_ignore"]
        #        from fews_utilities import xml_to_dict
        #        import regex as re

        location_sets = self.location_sets
        idmap_sections = self.idmap_sections

        loc_set_errors = {
            "locationId": [],
            "caw_code": [],
            "caw_name": [],
            "csv_file": [],
            "location_name": [],
            "type": [],
            "functie": [],
            "name_error": [],
            "caw_name_inconsistent": [],
            "missing_in_map": [],
            "missing_in_set": [],
            "missing_peilschaal": [],
            "missing_hbov": [],
            "missing_hben": [],
            "missing_hbovps": [],
            "missing_hbenps": [],
            "missing_hloc": [],
            "xy_not_same": [],
        }

        sets = {
            "waterstandlocaties": "WATERSTANDLOCATIES",
            "sublocaties": "KUNSTWERKEN",
            "hoofdlocaties": "KUNSTWERKEN",
        }

        for set_name, section_name in sets.items():
            self.logging.info(set_name)
            location_set = location_sets[set_name]
            location_gdf = location_set["gdf"]
            csv_file = self.fews_config.locationSets[location_set["id"]]["csvFile"][
                "file"
            ]
            int_locs = []

            for idmap in ["IdOPVLWATER", "IdOPVLWATER_HYMOS"]:
                for section in idmap_sections[idmap][section_name]:
                    int_locs += [
                        item["internalLocation"]
                        for item in xml_to_dict(
                            self.fews_config.IdMapFiles[idmap], **section
                        )["idMap"]["map"]
                    ]

            if set_name == "sublocaties":
                int_locs = [loc for loc in int_locs if not loc[-1] == "0"]
                par_gdf = location_sets["hoofdlocaties"]["gdf"]

            elif set_name == "hoofdlocaties":
                int_locs = [loc for loc in int_locs if loc[-1] == "0"]

            # idx, row = list(location_gdf.iterrows())[0]
            for idx, row in list(location_gdf.iterrows()):
                error = {
                    "name_error": False,
                    "caw_name_inconsistent": False,
                    "missing_in_map": False,
                    "type": "",
                    "functie": "",
                    "missing_in_set": False,
                    "missing_peilschaal": False,
                    "missing_hbov": False,
                    "missing_hben": False,
                    "missing_hbovps": False,
                    "missing_hbenps": False,
                    "missing_hloc": False,
                    "xy_not_same": False,
                }

                loc_id = row["LOC_ID"]
                caw_code = loc_id[2:-2]
                loc_name = row["LOC_NAME"]
                caw_name = ""

                if set_name == "sublocaties":

                    loc_functie = row["FUNCTIE"]
                    sub_type = row["TYPE"]

                    if sub_type in [
                        "afsluiter",
                        "debietmeter",
                        "krooshek",
                        "vispassage",
                    ]:
                        if not re.match(
                            f"[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*-{sub_type}", loc_name
                        ):
                            error["name_error"] = True

                    else:
                        if not re.match((
                            f"[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*-"
                            f"{sub_type}[0-9]_{loc_functie}"),
                            loc_name,
                        ):
                            error["name_error"] = True

                    if not error["name_error"]:
                        caw_name = re.match("([A-Z0-9 ]*)_", loc_name).group(1)
                        if not all(
                            location_gdf[
                                location_gdf["LOC_ID"].str.match(f"..{caw_code}")
                            ]["LOC_NAME"].str.match(f"({caw_name}_{caw_code}-K)")
                        ):
                            error["caw_name_inconsistent"] = True

                    if (
                        not row["HBOV"]
                        in location_sets["waterstandlocaties"]["gdf"]["LOC_ID"].values
                    ):
                        error["missing_hbov"] = True

                    if (
                        not row["HBEN"]
                        in location_sets["waterstandlocaties"]["gdf"]["LOC_ID"].values
                    ):
                        error["missing_hben"] = True

                    if (
                        not row["HBOVPS"]
                        in location_sets["peilschalen"]["gdf"]["LOC_ID"].values
                    ):
                        error["missing_hbovps"] = True

                    if (
                        not row["HBENPS"]
                        in location_sets["peilschalen"]["gdf"]["LOC_ID"].values
                    ):
                        error["missing_hbenps"] = True

                    if (
                        not row["PAR_ID"]
                        in location_sets["hoofdlocaties"]["gdf"]["LOC_ID"].values
                    ):
                        error["missing_hloc"] = True

                    else:
                        if not any(
                            [
                                re.match(loc, loc_id)
                                for loc in xy_ignore_df["internalLocation"]
                            ]
                        ):
                            if (
                                not par_gdf[par_gdf["LOC_ID"] == row["PAR_ID"]][
                                    "geometry"
                                ]
                                .values[0]
                                .equals(row["geometry"])
                            ):
                                error["xy_not_same"] = True

                    if any(error.values()):
                        error["type"] = sub_type
                        error["functie"] = loc_functie

                elif set_name == "hoofdlocaties":
                    if not re.match(f"[A-Z0-9 ]*_{caw_code}-K_[A-Z0-9 ]*", loc_name):
                        error["name_error"] = True

                elif set_name == "waterstandlocaties":
                    if not re.match(f"[A-Z0-9 ]*_{caw_code}-w_.*", loc_name):
                        error["name_error"] = True

                    if not error["name_error"]:
                        caw_name = re.match("([A-Z0-9 ]*)_", loc_name).group(1)
                        if not all(
                            location_gdf[
                                location_gdf["LOC_ID"].str.match(f"..{caw_code}")
                            ]["LOC_NAME"].str.match(f"({caw_name}_{caw_code}-w)")
                        ):
                            error["caw_name_inconsistent"] = True

                    if (
                        not row["PEILSCHAAL"]
                        in location_sets["peilschalen"]["gdf"]["LOC_ID"].values
                    ):
                        error["missing_peilschaal"] = True

                    if loc_id not in int_locs:
                        error["missing_in_map"] = True

                if any(error.values()):
                    loc_set_errors["locationId"].append(loc_id)
                    loc_set_errors["caw_name"].append(caw_name)
                    loc_set_errors["caw_code"].append(caw_code)
                    loc_set_errors["csv_file"].append(csv_file)
                    loc_set_errors["location_name"].append(loc_name)
                    for key, value in error.items():
                        loc_set_errors[key].append(value)

        self.consistency["locSet error"] = pd.DataFrame(loc_set_errors)
        # opname in samenvatting

        if len(self.consistency["locSet error"]) == 0:
            self.logging.info("no errors in locationSets")
        else:
            self.logging.warning(
                "{} errors in locationSets".format(
                    len(self.consistency["locSet error"])
                )
            )

    def write_excel(self):
        """Write consistency to excel."""
        consistency_xlsx = self.paths["consistency_xlsx"]
        consistency_out_xlsx = consistency_xlsx.parent.joinpath(
            f"{consistency_xlsx.stem}_uit.xlsx"
        )

        index = self.consistency["inhoudsopgave"]
        index.index = index["werkblad"]
        summary = {
            key: len(df)
            for key, df in self.consistency.items()
            if key not in self.fixed_sheets + ["inhoudsopgave"]
        }

        # read input xlsx and empty all warning sheets
        book = load_workbook(consistency_xlsx)
        for worksheet in book.worksheets:
            if worksheet.title not in self.fixed_sheets:
                book.remove(worksheet)

        # add summary
        worksheet = book.create_sheet("samenvatting", 1)
        worksheet.sheet_properties.tabColor = "92D050"
        worksheet.append(["controle", "aantal", "beschrijving"])
        for cell in worksheet["{}".format(worksheet.max_row)]:
            cell.font = Font(bold=True)

        for key, value in summary.items():
            worksheet.append([key, value, index.loc[key]["beschrijving"]])
            if (value > 0) and (key != "mpt"):
                worksheet[worksheet.max_row][1].fill = PatternFill(
                    fgColor="FF0000", fill_type="solid"
                )
            else:
                worksheet[worksheet.max_row][1].fill = PatternFill(
                    fgColor="92D050", fill_type="solid"
                )

        worksheet.column_dimensions["A"].width = 40
        worksheet.column_dimensions["C"].width = 100
        worksheet.auto_filter.ref = worksheet.dimensions

        # add warning sheets
        xls_writer = pd.ExcelWriter(consistency_out_xlsx, engine="openpyxl")
        xls_writer.book = book

        for sheet_name in summary.keys():
            df = self.consistency[sheet_name]
            if not df.empty:
                if df.index.name is None:
                    df.to_excel(xls_writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(xls_writer, sheet_name=sheet_name, index=True)
                worksheet = xls_writer.sheets[sheet_name]
                for col in worksheet.columns:
                    worksheet.column_dimensions[col[0].column_letter].width = 20
                worksheet.auto_filter.ref = worksheet.dimensions
                worksheet.freeze_panes = worksheet["B2"]
                if not df.empty:
                    if sheet_name == "mpt":
                        worksheet.sheet_properties.tabColor = "92D050"
                    else:
                        worksheet.sheet_properties.tabColor = "FF0000"

        xls_writer.book.active = xls_writer.book["samenvatting"]
        xls_writer.save()

    def write_csvs(self):
        """Write locationSets to CSV."""
        if "mpt" not in self.consistency.keys():
            self.hist_tags_to_mpt()
        mpt_df = self.consistency["mpt"]

        if self.waterstandloc is None:
            self._read_locs()

        date_threshold = mpt_df["ENDDATE"].max() - pd.Timedelta(weeks=26)

        location_sets = {
            key: value
            for key, value in self.location_sets.items()
            if value["id"]
            in ["OPVLWATER_HOOFDLOC", "OPVLWATER_WATERSTANDEN_AUTO", "OPVLWATER_SUBLOC"]
        }

        for key, value in location_sets.items():
            self.logging.info(f"writing CSV for set: {key}")
            gdf = value["gdf"]
            df = gdf.drop("geometry", axis=1)
            df[["START", "EIND"]] = df.apply(
                update_date, args=(mpt_df, date_threshold), axis=1, result_type="expand"
            )

            if value["id"] == "OPVLWATER_WATERSTANDEN_AUTO":
                grouper = self.mpt_hist_tags.groupby(["fews_locid"])
                df["HIST_TAG"] = df.apply(
                    update_histtag, args=[grouper], axis=1, result_type="expand"
                )

            elif value["id"] == "OPVLWATER_SUBLOC":
                grouper = df.groupby(["PAR_ID"])
                par_types_df = (
                    grouper["TYPE"]
                    .unique()
                    .apply(lambda x: sorted(x))
                    .transform(lambda x: "/".join(x))
                )

                df["PAR_ID"] = gdf["LOC_ID"].str[0:-1] + "0"
                df["ALLE_TYPES"] = df["PAR_ID"].apply(lambda x: par_types_df.loc[x])
                df[["HBOVPS", "HBENPS"]] = df.apply(
                    self._update_staff_gauge, axis=1, result_type="expand"
                )

            csv_file = self.paths["csv_out"].joinpath(
                self.fews_config.locationSets[value["id"]]["csvFile"]["file"]
            )
            if csv_file.suffix == "":
                csv_file = Path(f"{csv_file}.csv")
            df.to_csv(csv_file, index=False)
